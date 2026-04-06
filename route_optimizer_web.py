#!/usr/bin/env python3
"""
Innofoods -> McLane Warehouse Route Optimizer - Web Interface

Flask web application wrapping the route_optimizer.py CLI tool.
Run: python3 route_optimizer_web.py
Port: 8105
"""

import json
import sys
import os
from datetime import datetime, timedelta

# Ensure the route_optimizer module can be imported
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from flask import Flask, request, jsonify, render_template_string, Response

from route_optimizer import (
    WAREHOUSES,
    WH_BY_NAME,
    COORDS,
    ORIGIN_COORDS,
    RECEIVING_HOURS,
    _raw_schedule,
    BASE_COST,
    COST_PER_KM,
    MAX_STOPS,
    MAX_CAPACITY,
    AVG_SPEED_KMH,
    LOADING_TIME_H,
    UNLOAD_TIME_H,
    DEFAULT_MS_ORIGIN,
    DEFAULT_L1_ORIGIN,
    resolve_origin,
    optimize_group,
    load_cache,
    save_cache,
    generate_sample,
    reload_data,
    _load_warehouse_data,
    _save_warehouse_data,
)

app = Flask(__name__)

# ---------------------------------------------------------------------------
# DC code -> warehouse name lookup (built from warehouse data)
# ---------------------------------------------------------------------------
def _build_dc_map():
    """Build DC_CODE_MAP from warehouse data.
    If a DC code is used by multiple warehouses, create unique keys like 'ME|MS-WH', 'ME|L1-WH'.
    """
    # Count DC code usage
    dc_count = {}
    for wh in WAREHOUSES:
        dc = wh.get("dc_code", "")
        if dc:
            dc_count[dc] = dc_count.get(dc, 0) + 1

    m = {}
    for wh in WAREHOUSES:
        dc = wh.get("dc_code", "")
        if dc:
            if dc_count[dc] > 1:
                # Duplicate DC - use DC|group as key
                key = f"{dc}|{wh['group']}"
            else:
                key = dc
            m[key] = wh["name"]
    return m

def _reload_all():
    """Reload warehouse data and rebuild DC map."""
    global DC_CODE_MAP
    reload_data()
    DC_CODE_MAP = _build_dc_map()

DC_CODE_MAP = _build_dc_map()


def _resolve_dc(dc_code, po_label, warnings, ship_to=""):
    """Resolve a DC code to warehouse name, handling duplicates.
    If duplicate DC, try matching ship_to against warehouse names/addresses.
    Returns (warehouse_name, dc_key) or ("", "") if not found.
    """
    # Direct match (unique DC)
    if dc_code in DC_CODE_MAP:
        return DC_CODE_MAP[dc_code], dc_code

    # Check for duplicate DC keys (e.g., ME|MS-WH, ME|L1-WH)
    matches = {k: v for k, v in DC_CODE_MAP.items() if k.startswith(dc_code + "|")}
    if len(matches) == 1:
        key = list(matches.keys())[0]
        return matches[key], key
    elif len(matches) > 1:
        # Try to disambiguate using ship_to text
        if ship_to:
            ship_upper = ship_to.upper()
            for key, wh_name in matches.items():
                # Check if warehouse name or address appears in ship_to
                wh_data = WH_BY_NAME.get(wh_name, {})
                wh_addr = wh_data.get("address", "").upper() if wh_data else ""
                name_upper = wh_name.upper()
                # Match by name keywords
                name_words = [w for w in name_upper.split() if len(w) > 2]
                if any(w in ship_upper for w in name_words):
                    return wh_name, key
                # Match by address keywords
                if wh_addr:
                    addr_words = [w for w in wh_addr.split() if len(w) > 3]
                    if any(w in ship_upper for w in addr_words):
                        return wh_name, key

        # Still ambiguous - pick first and warn
        key = sorted(matches.keys())[0]
        groups = [k.split("|")[1] for k in matches.keys()]
        warnings.append(f'PO {po_label}: DC "{dc_code}" 가 여러 창고에 사용됨 ({", ".join(groups)}). 수동 선택 필요.')
        return matches[key], key

    warnings.append(f'PO {po_label}: DC 코드 "{dc_code}" 를 찾을 수 없습니다.')
    return "", ""


# ---------------------------------------------------------------------------
# Sample data - real PO data (Revised QTY = pallets)
# ---------------------------------------------------------------------------
SAMPLE_POS = [
    {"po_number": "MS10073402-01", "warehouse": "McLane Sun West AZ",            "due_date": "2026-03-24", "quantity": 24, "product_type": "DSKT", "inventory_available_date": "2026-03-11"},
    {"po_number": "MW10094380-01", "warehouse": "McLane Western CO",             "due_date": "2026-03-24", "quantity": 44, "product_type": "DSKT", "inventory_available_date": "2026-03-11"},
    {"po_number": "HP10067649-01", "warehouse": "McLane High Plains TX",         "due_date": "2026-03-24", "quantity": 14, "product_type": "DSKT", "inventory_available_date": "2026-03-11"},
    {"po_number": "NT10187696-01", "warehouse": "McLane North Texas TX",         "due_date": "2026-03-24", "quantity": 38, "product_type": "DSKT", "inventory_available_date": "2026-03-11"},
    {"po_number": "NW10089277-01", "warehouse": "McLane Northwest WA",           "due_date": "2026-03-24", "quantity": 4,  "product_type": "DSKT", "inventory_available_date": "2026-03-11"},
    {"po_number": "SZ10081334-01", "warehouse": "McLane Southern California CA", "due_date": "2026-03-24", "quantity": 20, "product_type": "DSKT", "inventory_available_date": "2026-03-11"},
    {"po_number": "SW10095010-01", "warehouse": "McLane Southwest TX",           "due_date": "2026-03-24", "quantity": 34, "product_type": "DSKT", "inventory_available_date": "2026-03-11"},
    {"po_number": "MN10098698-01", "warehouse": "McLane Minnesota MN",           "due_date": "2026-03-24", "quantity": 20, "product_type": "DSKT", "inventory_available_date": "2026-03-11"},
    {"po_number": "MP10072671-01", "warehouse": "McLane Pacific CA",             "due_date": "2026-03-24", "quantity": 12, "product_type": "DSKT", "inventory_available_date": "2026-03-11"},
    {"po_number": "NC10097501-01", "warehouse": "McLane Carolina NC",            "due_date": "2026-03-24", "quantity": 30, "product_type": "DSKT", "inventory_available_date": "2026-03-11"},
    {"po_number": "MZ10084952-01", "warehouse": "McLane Mid Atlantic VA",        "due_date": "2026-03-24", "quantity": 12, "product_type": "DSKT", "inventory_available_date": "2026-03-11"},
    {"po_number": "NE10085348-01", "warehouse": "McLane Northeast NY",           "due_date": "2026-03-24", "quantity": 14, "product_type": "DSKT", "inventory_available_date": "2026-03-11"},
    {"po_number": "PA10086419-01", "warehouse": "McLane Pennsylvania PA",        "due_date": "2026-03-24", "quantity": 20, "product_type": "DSKT", "inventory_available_date": "2026-03-11"},
    {"po_number": "MY10075692-01", "warehouse": "McLane Concord NH",             "due_date": "2026-03-24", "quantity": 6,  "product_type": "DSKT", "inventory_available_date": "2026-03-11"},
    {"po_number": "MD10087384-01", "warehouse": "McLane Dothan AL",              "due_date": "2026-03-24", "quantity": 10, "product_type": "DSKT", "inventory_available_date": "2026-03-11"},
    {"po_number": "SO10081532-01", "warehouse": "McLane Brookhaven MS",          "due_date": "2026-03-24", "quantity": 24, "product_type": "DSKT", "inventory_available_date": "2026-03-11"},
]


# ---------------------------------------------------------------------------
# Helper: run optimization
# ---------------------------------------------------------------------------
def run_optimization(pos_data, ms_origin_str, l1_origin_str, balance_weight=0.0):
    """Run the optimizer and return structured results."""
    errors = []

    # Validate
    required_fields = {"po_number", "warehouse", "due_date", "quantity"}
    for po in pos_data:
        missing = required_fields - set(po.keys())
        if missing:
            errors.append(f"PO {po.get('po_number', '?')} missing fields: {missing}")
        if po.get("warehouse") and po["warehouse"] not in WH_BY_NAME:
            errors.append(f"Unknown warehouse '{po['warehouse']}' in PO {po.get('po_number', '?')}")
        # Ensure quantity is int
        try:
            po["quantity"] = int(po["quantity"])
        except (ValueError, TypeError):
            errors.append(f"PO {po.get('po_number', '?')}: invalid quantity")
        # Set inventory_available_date to 1 week before due_date
        if po.get("due_date"):
            try:
                dd = datetime.strptime(po["due_date"][:10], "%Y-%m-%d")
                po["inventory_available_date"] = (dd - timedelta(days=7)).strftime("%Y-%m-%d")
            except ValueError:
                po["inventory_available_date"] = datetime.now().strftime("%Y-%m-%d")
        elif not po.get("inventory_available_date"):
            po["inventory_available_date"] = datetime.now().strftime("%Y-%m-%d")

    if errors:
        return {"success": False, "errors": errors}

    # Split by group
    ms_pos = [po for po in pos_data if WH_BY_NAME[po["warehouse"]]["group"] == "MS-WH"]
    l1_pos = [po for po in pos_data if WH_BY_NAME[po["warehouse"]]["group"] == "L1-WH"]

    ms_origin = resolve_origin(ms_origin_str)
    l1_origin = resolve_origin(l1_origin_str)

    cache = load_cache()

    ms_result = optimize_group("MS-WH", ms_pos, ms_origin, ms_origin_str, cache, balance_weight)
    l1_result = optimize_group("L1-WH", l1_pos, l1_origin, l1_origin_str, cache, balance_weight)

    save_cache(cache)

    grand_total = ms_result.get("total_cost", 0) + l1_result.get("total_cost", 0)
    total_routes = ms_result.get("total_routes", 0) + l1_result.get("total_routes", 0)
    total_infeasible = len(ms_result.get("infeasible", [])) + len(l1_result.get("infeasible", []))

    # Build coordinates map for map display
    coords_map = {}
    for wh in WAREHOUSES:
        if wh["name"] in COORDS:
            c = COORDS[wh["name"]]
            coords_map[wh["name"]] = {"lat": c[0], "lon": c[1]}
    # Add origins
    coords_map["__ms_origin__"] = {"lat": ms_origin[0], "lon": ms_origin[1], "label": ms_origin_str}
    coords_map["__l1_origin__"] = {"lat": l1_origin[0], "lon": l1_origin[1], "label": l1_origin_str}

    # Build receiving schedule lookup: code -> human-readable string
    recv_schedules = {}
    for code, windows in _raw_schedule.items():
        parts = []
        for day_spec, start_str, end_str in windows:
            if start_str.upper() in ("CLOSE", "CLOSED"):
                parts.append(f"{day_spec} CLOSED")
            else:
                parts.append(f"{day_spec} {start_str}-{end_str}")
        recv_schedules[code] = " / ".join(parts)

    return {
        "success": True,
        "ms_result": ms_result,
        "l1_result": l1_result,
        "coords": coords_map,
        "receiving_schedules": recv_schedules,
        "summary": {
            "total_pos": len(pos_data),
            "ms_pos": len(ms_pos),
            "l1_pos": len(l1_pos),
            "total_routes": total_routes,
            "grand_total": round(grand_total, 2),
            "total_infeasible": total_infeasible,
        },
        "assumptions": {
            "base_cost": BASE_COST,
            "cost_per_km": COST_PER_KM,
            "max_stops": MAX_STOPS,
            "max_capacity": MAX_CAPACITY,
            "avg_speed_kmh": AVG_SPEED_KMH,
            "loading_time_h": LOADING_TIME_H,
            "unload_time_min": UNLOAD_TIME_H * 60,
        },
    }


# ---------------------------------------------------------------------------
# HTML Templates
# ---------------------------------------------------------------------------

MAIN_PAGE = r"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Route Optimizer - Innofoods</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css" rel="stylesheet">
<link href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" rel="stylesheet">
<style>
  :root { --cj-blue: #003d7a;
  .modal-backdrop { z-index: 10005 !important; }
  .modal { z-index: 10010 !important; } --cj-light: #e8f0fe; --cj-accent: #0056b3; }
  body { background: linear-gradient(135deg, #f0f2f5 0%, #e2e8f0 100%); font-size: 0.92rem; min-height: 100vh; }
  .navbar { background: linear-gradient(90deg, #001f3f 0%, var(--cj-blue) 50%, #004e9a 100%) !important; box-shadow: 0 2px 12px rgba(0,0,0,.2); }
  .navbar-brand { font-size: 1.25rem; letter-spacing: 0.5px; }
  .card { border: none; box-shadow: 0 2px 12px rgba(0,0,0,.06); border-radius: 12px; transition: box-shadow 0.2s; }
  .card:hover { box-shadow: 0 4px 20px rgba(0,0,0,.1); }
  .table-striped>tbody>tr:nth-of-type(odd)>* { background-color: var(--cj-light); }
  .table-dark { background: var(--cj-blue); }
  .btn-cj { background: linear-gradient(135deg, var(--cj-blue), var(--cj-accent)); color: #fff; border: none; box-shadow: 0 2px 8px rgba(0,61,122,.3); transition: all 0.2s; }
  .btn-cj:hover { background: linear-gradient(135deg, #00274d, var(--cj-blue)); color: #fff; box-shadow: 0 4px 16px rgba(0,61,122,.4); transform: translateY(-1px); }
  .hero-section { text-align: center; padding: 2rem 0 1rem; }
  .hero-section h2 { color: var(--cj-blue); font-weight: 700; margin-bottom: 0.25rem; }
  .hero-section p { color: #6c757d; font-size: 0.95rem; }
  .toolbar-card { background: linear-gradient(135deg, #fff 0%, #f8fafc 100%); }
  .toolbar-card .btn { border-radius: 8px; font-size: 0.82rem; }
  #loading-overlay {
    display: none; position: fixed; inset: 0; background: rgba(0,15,30,.55);
    z-index: 9999; align-items: center; justify-content: center;
    backdrop-filter: blur(4px);
  }
  #loading-overlay.active { display: flex; }
  .spinner-box { background: #fff; border-radius: 16px; padding: 2.5rem 3rem; text-align: center; box-shadow: 0 8px 32px rgba(0,0,0,.15); }
  .dest-hidden { display: none !important; }
  .optimize-btn-wrapper { padding: 1rem 0 0.5rem; }
  .optimize-btn-wrapper .btn { font-size: 1.1rem; border-radius: 12px; padding: 0.75rem 3rem; }
  @media print {
    .no-print { display: none !important; }
    body { background: #fff; font-size: 0.82rem; }
    .card { box-shadow: none; border: 1px solid #ddd; border-radius: 0; }
  }
</style>
</head>
<body>

<!-- Loading overlay -->
<div id="loading-overlay">
  <div class="spinner-box">
    <div class="spinner-border text-primary mb-3" style="width:3rem;height:3rem;" role="status"></div>
    <p class="mb-0 fw-semibold">Optimizing routes...<br><small class="text-muted">OSRM distance calculation may take a few seconds</small></p>
  </div>
</div>

<nav class="navbar navbar-dark mb-0 no-print">
  <div class="container-fluid">
    <span class="navbar-brand mb-0"><i class="bi bi-truck"></i> Innofoods Route Optimizer</span>
    <div class="d-flex gap-2">
      <a href="/warehouses" class="btn btn-outline-light btn-sm"><i class="bi bi-building"></i> Warehouse 관리</a>
      <a href="/learned" class="btn btn-outline-light btn-sm"><i class="bi bi-mortarboard"></i> 학습 이력</a>
      <a href="/cache" class="btn btn-outline-light btn-sm"><i class="bi bi-database"></i> Distance Cache</a>
    </div>
  </div>
</nav>

<div class="container-fluid" style="max-width:1400px;">

  <!-- Hero -->
  <div class="hero-section no-print">
    <h2><i class="bi bi-geo-alt-fill"></i> Delivery Route Optimizer</h2>
    <p>McLane Warehouse Delivery Route Optimization System</p>
  </div>

  <!-- Input Section -->
  <div id="input-section">
    <!-- Toolbar -->
    <div class="card toolbar-card p-3 mb-3">
      <div class="d-flex justify-content-between align-items-center">
        <h6 class="mb-0"><i class="bi bi-table"></i> PO Data (주문 데이터)</h6>
        <div class="no-print d-flex gap-1 flex-wrap align-items-center">
          <button class="btn btn-sm btn-outline-secondary" onclick="addRow()"><i class="bi bi-plus-lg"></i> Add Row</button>
          <a class="btn btn-sm btn-outline-success" href="/api/template.xlsx"><i class="bi bi-file-earmark-excel"></i> Template</a>
          <button class="btn btn-sm btn-outline-warning" onclick="document.getElementById('excel-upload').click()"><i class="bi bi-upload"></i> Upload Excel</button>
          <input type="file" id="excel-upload" accept=".xlsx,.xls" style="display:none" onchange="uploadExcel(this)">
          <button class="btn btn-sm btn-outline-danger" onclick="clearAll()"><i class="bi bi-trash"></i> Clear</button>
          <span class="ms-2">|</span>
          <span class="ms-2 d-flex align-items-center gap-1" title="0 = 비용 최소화, 높을수록 루트 균등 배분 우선">
            <label class="form-label mb-0 small text-nowrap" for="balance-slider">Balance:</label>
            <input type="range" class="form-range" id="balance-slider" min="0" max="1" step="0.1" value="0.3" style="width:80px;">
            <span class="badge bg-secondary" id="balance-val">0.3</span>
          </span>
          <button class="btn btn-sm btn-cj ms-2" onclick="optimize()"><i class="bi bi-cpu"></i> Optimize Routes (최적화 실행)</button>
        </div>
      </div>
    </div>

    <!-- PO Table -->
    <div class="card p-3 mb-3">
      <div></div>
      <div class="table-responsive">
        <table class="table table-sm table-striped align-middle mb-0" id="po-table">
          <thead class="table-dark">
            <tr>
              <th style="width:35px;"></th>
              <th style="width:280px;">DC</th>
              <th style="width:150px;">PO #</th>
              <th style="width:80px;">Item</th>
              <th style="width:100px;">Revised QTY</th>
              <th style="width:130px;">REQ DELIVERY</th>
              <th style="width:130px;">Inv. Available</th>
            </tr>
          </thead>
          <tbody id="po-tbody"></tbody>
        </table>
      </div>
    </div>

  </div>

  <!-- Results Section (hidden initially) -->
  <div id="results-section" style="display:none;"></div>

</div>

<!-- AI Chat Panel -->
<div id="ai-chat-panel" style="display:none;position:fixed;right:16px;bottom:16px;width:420px;max-height:70vh;z-index:9998;background:#fff;border-radius:12px;box-shadow:0 4px 24px rgba(0,0,0,.2);overflow:hidden;">
  <div style="background:linear-gradient(90deg,#001f3f,#003d7a);color:#fff;padding:8px 16px;display:flex;align-items:center;justify-content:space-between;">
    <span><i class="bi bi-robot"></i> AI Route Analyst</span>
    <button class="btn btn-sm btn-outline-light py-0" onclick="toggleAIChat()"><i class="bi bi-x-lg"></i></button>
  </div>
  <div id="ai-chat-messages" style="height:calc(70vh - 110px);overflow-y:auto;padding:12px;font-size:0.85rem;"></div>
  <div style="padding:8px;border-top:1px solid #eee;display:flex;gap:6px;">
    <input type="text" id="ai-chat-input" class="form-control form-control-sm" placeholder="질문하세요... (예: 가장 비싼 루트는?)" onkeydown="if(event.key==='Enter')sendAIChat()">
    <button class="btn btn-sm btn-cj" onclick="sendAIChat()"><i class="bi bi-send"></i></button>
  </div>
</div>

<!-- Route Detail Modal (bottom sheet) -->
<div class="modal fade" id="routeDetailModal" tabindex="-1" style="z-index:10010;">
  <div class="modal-dialog modal-xl modal-dialog-end" style="position:fixed;bottom:0;left:50%;transform:translateX(-50%);margin:0;max-width:95vw;width:95vw;">
    <div class="modal-content" style="border-radius:12px 12px 0 0;box-shadow:0 -4px 20px rgba(0,0,0,.15);">
      <div class="modal-header py-2" style="background:linear-gradient(90deg,#001f3f,#003d7a);color:#fff;border-radius:12px 12px 0 0;">
        <h6 class="modal-title"><i class="bi bi-signpost-2"></i> Route Detail</h6>
        <button type="button" class="btn-close btn-close-white" data-bs-dismiss="modal"></button>
      </div>
      <div class="modal-body py-2" id="routeModalBody" style="max-height:45vh;overflow-y:auto;"></div>
    </div>
  </div>
</div>

<!-- Error Modal -->
<div class="modal fade" id="errorModal" tabindex="-1" style="z-index:10010;">
  <div class="modal-dialog modal-dialog-centered">
    <div class="modal-content">
      <div class="modal-header bg-danger text-white py-2">
        <h6 class="modal-title"><i class="bi bi-exclamation-triangle"></i> Error</h6>
        <button type="button" class="btn-close btn-close-white" data-bs-dismiss="modal"></button>
      </div>
      <div class="modal-body" id="errorModalBody" style="white-space:pre-wrap;"></div>
      <div class="modal-footer py-1">
        <button class="btn btn-sm btn-secondary" data-bs-dismiss="modal">닫기</button>
      </div>
    </div>
  </div>
</div>

<!-- Import Modal -->
<div class="modal fade" id="importModal" tabindex="-1">
  <div class="modal-dialog modal-lg">
    <div class="modal-content">
      <div class="modal-header"><h5 class="modal-title">Import JSON</h5><button type="button" class="btn-close" data-bs-dismiss="modal"></button></div>
      <div class="modal-body">
        <p class="text-muted small">Paste a JSON array of PO objects. Required: dc (or warehouse), po_number, due_date, quantity</p>
        <textarea class="form-control" id="import-json" rows="12" placeholder='[{"dc":"MS","po_number":"MS10073402-01","product_type":"DSKT","quantity":24,"due_date":"2026-03-24"}]'></textarea>
      </div>
      <div class="modal-footer">
        <button class="btn btn-secondary" data-bs-dismiss="modal">Cancel</button>
        <button class="btn btn-cj" onclick="doImport()">Import</button>
      </div>
    </div>
  </div>
</div>

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script>
document.getElementById('balance-slider').addEventListener('input', function() {
  document.getElementById('balance-val').textContent = this.value;
});
const WAREHOUSES = {{ warehouses_json|safe }};
const SAMPLE = {{ sample_json|safe }};
const DC_MAP = {{ dc_map_json|safe }};
const DEFAULT_MS = '{{ default_ms }}';
const DEFAULT_L1 = '{{ default_l1 }}';

// Build sorted DC list for dropdown
const DC_LIST = Object.keys(DC_MAP).sort();

function dcOptions(selected) {
  var html = '<option value="">--</option>';
  for (var i = 0; i < DC_LIST.length; i++) {
    var key = DC_LIST[i];
    var wh = DC_MAP[key];
    var sel = (key === selected) ? 'selected' : '';
    // Display: if key has '|' (duplicate DC), show "ME (MS-WH) - Name", else "ME - Name"
    var label = key.includes('|') ? key.replace('|', ' (') + ') - ' + wh : key + ' - ' + wh;
    html += '<option value="' + key + '" ' + sel + '>' + label + '</option>';
  }
  return html;
}

function addRow(data) {
  var d = data || {};
  var tbody = document.getElementById('po-tbody');
  var tr = document.createElement('tr');
  tr.innerHTML =
    '<td><button class="btn btn-sm btn-outline-danger py-0 px-1" onclick="this.closest(\'tr\').remove()"><i class="bi bi-x"></i></button></td>' +
    '<td><select class="form-select form-select-sm po-field" data-field="dc">' + dcOptions(d.dc || '') + '</select></td>' +
    '<td><input type="text" class="form-control form-control-sm po-field" data-field="po_number" value="' + (d.po_number || '') + '"></td>' +
    '<td><input type="text" class="form-control form-control-sm po-field" data-field="product_type" value="' + (d.product_type || '') + '"></td>' +
    '<td><input type="number" class="form-control form-control-sm po-field" data-field="quantity" value="' + (d.quantity || '') + '" min="1"></td>' +
    '<td><input type="date" class="form-control form-control-sm po-field" data-field="due_date" value="' + (d.due_date || '') + '"></td>' +
    '<td><input type="date" class="form-control form-control-sm po-field" data-field="inventory_available_date" value="' + (d.inventory_available_date || '') + '"></td>';
  tbody.appendChild(tr);
}

function clearAll() {
  document.getElementById('po-tbody').innerHTML = '';
}

function loadSample() {
  clearAll();
  SAMPLE.forEach(d => addRow(d));
}

function showImportModal() {
  new bootstrap.Modal(document.getElementById('importModal')).show();
}

function doImport() {
  try {
    const data = JSON.parse(document.getElementById('import-json').value);
    if (!Array.isArray(data)) throw new Error('Must be an array');
    clearAll();
    data.forEach(d => addRow(d));
    bootstrap.Modal.getInstance(document.getElementById('importModal')).hide();
  } catch(e) {
    alert('Invalid JSON: ' + e.message);
  }
}

function gatherPOs() {
  var rows = document.querySelectorAll('#po-tbody tr');
  var pos = [];
  for (var i = 0; i < rows.length; i++) {
    var tr = rows[i];
    var po = {};
    var fields = tr.querySelectorAll('.po-field');
    for (var j = 0; j < fields.length; j++) {
      var el = fields[j];
      var v = el.value.trim();
      if (el.dataset.field === 'quantity') v = parseInt(v) || 0;
      po[el.dataset.field] = v;
    }
    // Resolve DC code to warehouse name
    if (po.dc && DC_MAP[po.dc]) {
      po.warehouse = DC_MAP[po.dc];
    }
    if (po.due_date) {
      var dd = new Date(po.due_date);
      dd.setDate(dd.getDate() - 7);
      po.inventory_available_date = dd.toISOString().slice(0, 10);
    } else if (!po.inventory_available_date) {
      po.inventory_available_date = new Date().toISOString().slice(0, 10);
    }
    if (po.po_number) pos.push(po);
  }
  return pos;
}

async function optimize() {
  var pos = gatherPOs();
  if (pos.length === 0) { alert('PO 데이터를 추가해주세요.'); return; }

  var payload = {
    pos: pos,
    ms_origin: '{{ default_ms }}',
    l1_origin: '{{ default_l1 }}',
    balance_weight: parseFloat(document.getElementById('balance-slider').value),
  };

  document.getElementById('loading-overlay').classList.add('active');

  try {
    const resp = await fetch('/api/optimize', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify(payload),
    });
    const result = await resp.json();
    document.getElementById('loading-overlay').classList.remove('active');

    if (!result.success) {
      alert('Errors:\n' + (result.errors||[]).join('\n'));
      return;
    }
    window._lastResult = result;
    window._lastPayload = payload;
    window._originalTotal = null;  // reset for fresh comparison
    window._originalRoutes = null;
    window._prevL1Cost = 0; window._prevMSCost = 0; window._initL1Cost = 0; window._initMSCost = 0; window._initTotalCost = 0;  // reset cost tracking
    if (window._lastResult) delete window._lastResult._detached;  // clear detached
    renderResults(result, payload);
    // Auto-run schedule after results load
    setTimeout(function() { autoSchedule(); }, 500);
  } catch(e) {
    document.getElementById('loading-overlay').classList.remove('active');
    alert('Request failed: ' + e.message);
  }
}

async function recalcRoute(input) {
  var gi = parseInt(input.dataset.routeGroup);
  var ri = parseInt(input.dataset.routeIdx);
  var newPickup = input.value.replace('T', ' ');
  var result = window._lastResult;
  if (!result) return;

  var gr = [result.ms_result, result.l1_result][gi];
  if (!gr || !gr.routes[ri]) return;
  var rt = gr.routes[ri];

  // Gather POs for this route and recalculate server-side
  try {
    const resp = await fetchTimeout('/api/recalc-route', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({
        po_numbers: rt.po_numbers,
        warehouses: rt.warehouses,
        quantities: rt.schedule.map(s => s.quantity),
        due_dates: rt.schedule.map(s => s.due_date),
        pickup_time: newPickup,
        group: gr.group,
      }),
    });
    const data = await resp.json();
    if (data.success) {
      // Update route in result
      gr.routes[ri].schedule = data.route.schedule;
      gr.routes[ri].departure = newPickup;
      gr.routes[ri].distance_km = data.route.distance_km;
      gr.routes[ri].cost = data.route.cost;
      renderResults(result, window._lastPayload);
    } else {
      alert('Recalc failed: ' + (data.error || 'unknown'));
    }
  } catch(e) {
    alert('Recalc error: ' + e.message);
  }
}

async function mergeSelected() {
  var result = window._lastResult;
  if (!result) return;

  var checks = document.querySelectorAll('.po-merge-chk:checked');
  if (checks.length < 2) { alert('합칠 PO를 2개 이상 선택하세요.'); return; }

  var grList = [result.ms_result, result.l1_result];

  // Collect selected POs and track which routes they came from
  var groups = new Set();
  var selectedPOs = [];
  var affectedRoutes = new Map(); // gi_ri -> {gi, ri}
  checks.forEach(chk => {
    var gi = parseInt(chk.dataset.group);
    var ri = parseInt(chk.dataset.route);
    groups.add(gi);
    selectedPOs.push({
      gi: gi, ri: ri, si: parseInt(chk.dataset.stop),
      po: chk.dataset.po, wh: chk.dataset.wh,
      qty: parseInt(chk.dataset.qty), due: chk.dataset.due
    });
    affectedRoutes.set(gi+'_'+ri, {gi: gi, ri: ri});
  });

  // Allow cross-group merge

  var totalUnits = selectedPOs.reduce((s, p) => s + p.qty, 0);
  if (totalUnits > 60) {
    if (!confirm(`합산 ${totalUnits} units은 최대 용량(60)을 초과합니다. 계속?`)) return;
  }

  var gi = selectedPOs[0].gi;
  var gr = grList[gi];

  // Find earliest pickup from affected routes
  var pickups = [];
  affectedRoutes.forEach(ar => { pickups.push(gr.routes[ar.ri].departure); });
  var pickup = pickups.sort()[0];

  // Call server to calculate merged route
  try {
    const resp = await fetchTimeout('/api/recalc-route', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({
        po_numbers: selectedPOs.map(p => p.po),
        warehouses: selectedPOs.map(p => p.wh),
        quantities: selectedPOs.map(p => p.qty),
        due_dates: selectedPOs.map(p => p.due),
        pickup_time: pickup,
        group: gr.group,
      }),
    });
    const data = await resp.json();
    if (!data.success) { showError('Merge failed: ' + (data.error || '')); return; }

    // Remove selected POs from their original routes
    // Build set of selected PO numbers per route
    var removals = {}; // ri -> set of po_numbers
    selectedPOs.forEach(p => {
      if (!removals[p.ri]) removals[p.ri] = new Set();
      removals[p.ri].add(p.po);
    });

    // Process each affected route: remove POs, recalc or delete
    var routesToDelete = [];
    for (var riStr in removals) {
      var ri = parseInt(riStr);
      var rt = gr.routes[ri];
      var remaining = rt.schedule.filter(st => !removals[ri].has(st.po_number));
      if (remaining.length === 0) {
        routesToDelete.push(ri);
      } else {
        // Recalculate route with remaining POs
        const resp2 = await fetchTimeout('/api/recalc-route', {
          method: 'POST',
          headers: {'Content-Type': 'application/json'},
          body: JSON.stringify({
            po_numbers: remaining.map(s => s.po_number),
            warehouses: remaining.map(s => s.warehouse),
            quantities: remaining.map(s => s.quantity),
            due_dates: remaining.map(s => s.due_date),
            pickup_time: rt.departure,
            group: gr.group,
          }),
        });
        const d2 = await resp2.json();
        if (d2.success) {
          gr.routes[ri].schedule = d2.route.schedule;
          gr.routes[ri].distance_km = d2.route.distance_km;
          gr.routes[ri].cost = d2.route.cost;
          gr.routes[ri].total_units = remaining.reduce((s,st) => s + st.quantity, 0);
          gr.routes[ri].po_numbers = remaining.map(s => s.po_number);
          gr.routes[ri].warehouses = remaining.map(s => s.warehouse);
        }
      }
    }

    // Delete empty routes (reverse order)
    routesToDelete.sort((a,b) => b-a).forEach(ri => gr.routes.splice(ri, 1));

    // Add merged route
    var merged = data.route;
    merged.po_numbers = selectedPOs.map(p => p.po);
    merged.warehouses = selectedPOs.map(p => p.wh);
    merged.departure = pickup;
    merged.total_units = totalUnits;
    gr.routes.push(merged);

    // Recalc totals
    gr.total_cost = gr.routes.reduce((s, r) => s + r.cost, 0);
    gr.total_routes = gr.routes.length;

    renderResults(result, window._lastPayload);
  } catch(e) {
    showError('Merge error: ' + e.message);
  }
}

function renderResults(result, payload) {
  const s = result.summary;
  const a = result.assumptions;

  let html = `
  <div class="no-print mb-3 d-flex gap-2 flex-wrap align-items-center">
    <button class="btn btn-outline-secondary" onclick="backToInput()"><i class="bi bi-arrow-left"></i> Back (수정하기)</button>

    <button class="btn btn-outline-success" onclick="exportJSON()"><i class="bi bi-download"></i> Export JSON</button>
    <button class="btn btn-outline-primary" onclick="printMap()"><i class="bi bi-map"></i> 지도 인쇄</button>
    <button class="btn btn-outline-secondary" onclick="window.print()"><i class="bi bi-printer"></i> 전체 인쇄</button>
    <button class="btn btn-outline-success" onclick="autoSchedule()"><i class="bi bi-calendar-check"></i> 스케줄 자동결정</button>
    <button class="btn btn-outline-danger" onclick="learnResult()"><i class="bi bi-mortarboard"></i> 이 결과 학습</button>
    <button class="btn btn-outline-info ms-auto" onclick="toggleAIChat()"><i class="bi bi-robot"></i> AI 분석</button>
  </div>

  <!-- Tabs -->
  <ul class="nav nav-tabs mb-3 no-print" id="resultTabs" role="tablist">
    <li class="nav-item"><a class="nav-link active" data-bs-toggle="tab" href="#tab-routes" role="tab"><i class="bi bi-signpost-2"></i> Route View</a></li>
    <li class="nav-item"><a class="nav-link" data-bs-toggle="tab" href="#tab-loadplan" role="tab"><i class="bi bi-table"></i> Load Plan</a></li>
  </ul>
  <div class="tab-content">
  <div class="tab-pane fade show active" id="tab-routes" role="tabpanel">

  <!-- Route Map -->
  <div class="card p-3 mb-3">
    <h6><i class="bi bi-map"></i> Route Map (경로 지도) <small class="text-muted">- 마커 클릭으로 선택, Split/Merge 가능</small></h6>
    <div style="position:relative;">
      <div id="route-map" style="height:75vh;min-height:500px;border-radius:8px;"></div>
      <div id="map-controls" style="position:absolute;bottom:30px;left:50px;z-index:999;display:flex;gap:6px;">
        <button class="btn btn-sm btn-warning" onclick="mapMerge()" title="합치기 (F1)"><i class="bi bi-arrows-collapse"></i> 합치기 <kbd>A</kbd></button>
        <button class="btn btn-sm btn-danger" onclick="cutSelectedSegment()" title="끊기 (F3)"><i class="bi bi-scissors"></i> 끊기 <kbd>D</kbd></button>

        <span class="badge bg-primary align-self-center" id="map-sel-count-top" style="display:none;">0</span>
        <button class="btn btn-sm btn-dark ms-2" onclick="toggleFullscreen()" title="전체화면 (F4)"><i class="bi bi-arrows-fullscreen" id="fs-icon"></i> <kbd>F4</kbd></button>
      </div>
      <div id="map-cost-l1" style="position:absolute;top:10px;left:50px;z-index:999;background:rgba(52,152,219,0.9);color:#fff;padding:6px 14px;border-radius:8px;font-weight:bold;font-size:14px;display:none;">L1: $0</div>
      <div id="map-cost-total" style="position:absolute;top:10px;left:50%;transform:translateX(-50%);z-index:999;background:rgba(0,0,0,0.8);color:#fff;padding:6px 14px;border-radius:8px;font-weight:bold;font-size:15px;display:none;">Total: $0</div>
      <div id="map-cost-ms" style="position:absolute;top:10px;right:10px;z-index:999;background:rgba(231,76,60,0.9);color:#fff;padding:6px 14px;border-radius:8px;font-weight:bold;font-size:14px;display:none;">MS: $0</div>
    </div>
    <div class="mt-2 small text-muted"><i class="bi bi-info-circle"></i> <b>Split:</b> 선(구간)을 클릭하면 그 지점에서 루트 분리 &nbsp;|&nbsp; <b>Merge:</b> 서로 다른 루트의 마커를 각각 하나씩 클릭 후 합치기 (루트 전체가 합쳐짐)</div>
    <div id="map-action-bar" style="display:none;" class="mt-1 p-2 bg-light border rounded d-flex align-items-center gap-2">
      <span class="badge bg-primary" id="map-sel-count">0</span> <span id="seg-selected" style="display:none;" class="badge bg-danger ms-1">선 선택됨</span> <span class="small"> <kbd>A</kbd> 합치기 <kbd>F2</kbd> Clear <kbd>D</kbd> 끊기</span>
    </div>
  </div>

  <!-- Assumptions -->
  <div class="card p-3 mb-3">
    <h6><i class="bi bi-info-circle"></i> Assumptions (가정)</h6>
    <div class="row small">
      <div class="col-md-6">
        <ul class="mb-0">
          <li>Base cost per truck: <strong>$${a.base_cost}</strong></li>
          <li>Cost per km: <strong>$${a.cost_per_km}</strong></li>
          <li>Max stops per truck: <strong>${a.max_stops}</strong></li>
          <li>Max capacity per truck: <strong>${a.max_capacity} units</strong></li>
        </ul>
      </div>
      <div class="col-md-6">
        <ul class="mb-0">
          <li>Average speed: <strong>${a.avg_speed_kmh} km/h</strong></li>
          <li>Loading time at origin: <strong>${a.loading_time_h} hour(s)</strong></li>
          <li>Unloading time per stop: <strong>${a.unload_time_min} min</strong></li>
          <li>MS-WH / L1-WH groups never mixed on same truck</li>
        </ul>
      </div>
    </div>
  </div>`;

  // Render each group
  [result.ms_result, result.l1_result].forEach((gr, gi) => {
    if (!gr || (gr.routes.length === 0 && gr.infeasible.length === 0)) return;
    const isMS = gr.group === 'MS-WH';
    const badge = isMS ? 'bg-danger' : 'bg-primary';

    html += `
    <div class="card p-3 mb-3">
      <h5><span class="badge ${badge} me-2">${gr.group}</span> Origin: ${gr.origin || 'N/A'}
        <span class="float-end fs-6">Routes: ${gr.routes.length} | Total: <span class="text-success fw-bold">$${gr.total_cost.toLocaleString()}</span></span>
      </h5>
      <hr>`;

    gr.routes.forEach((rt, ri) => {
      var distCost = (rt.cost - a.base_cost).toFixed(2);
      html += `
      <div id="route-detail-${gi}-${ri}"></div>
      <h6 class="mt-3">
        ${rt.schedule.map(s => s.dc_code).join(" → ")} <span class="badge bg-info">${rt.total_units} units</span>
        <span class="badge bg-secondary ms-2">${rt.po_numbers.join(', ')}</span>
        <span class="float-end text-success">$${rt.cost.toLocaleString()}</span>
      </h6>
      <div class="row small mb-1 align-items-center">
        <div class="col-3"><strong>Pickup:</strong> <input type="datetime-local" class="form-control form-control-sm d-inline-block" style="width:200px;" value="${rt.departure.replace(' ','T')}" data-route-group="${gi}" data-route-idx="${ri}" onchange="recalcRoute(this)"> <small class="text-muted">${addDayName(rt.departure).match(/\((\w+)\)/)?.[1]||''}</small></div>
        <div class="col-3"><strong>Distance:</strong> ${rt.distance_km.toLocaleString()} km</div>
        <div class="col-3"><strong>Units:</strong> ${rt.total_units}</div>
        <div class="col-3"><strong>Base:</strong> $${a.base_cost} + <strong>Dist:</strong> $${distCost}</div>
      </div>
      <table class="table table-sm table-bordered table-striped small mb-3">
        <thead class="table-light">
          <tr><th style="width:30px;"></th><th>Stop</th><th>DC</th><th>Warehouse</th><th>PO</th><th>Qty</th><th>Seg km</th><th>Hours</th><th>마지노 출발</th><th>Arrival</th><th>Adj Arrival</th><th>대기</th><th>Due Date</th><th>Operating Hours (운영시간)</th></tr>
        </thead>
        <tbody>`;
      rt.schedule.forEach(st => {
        const isLate = st.adjusted_arrival > st.due_date + ' 23:59';
        const cls = isLate ? 'table-danger' : '';
        var opHrs = (result.receiving_schedules && st.receiving_code) ? (result.receiving_schedules[st.receiving_code] || st.receiving_code) : '';
        var segHrs = st.segment_hours || (st.segment_km / 80).toFixed(1);
        var restInfo = (st.rest_hours && st.rest_hours > 0) ? '<br><small class="text-muted">+' + st.rest_hours + 'h rest</small>' : '';
        html += `<tr class="${cls}"><td><input type="checkbox" class="form-check-input po-merge-chk" data-group="${gi}" data-route="${ri}" data-stop="${st.stop-1}" data-po="${st.po_number}" data-wh="${st.warehouse}" data-qty="${st.quantity}" data-due="${st.due_date}"></td><td>${st.stop}</td><td><span class="badge bg-secondary">${st.dc_code||''}</span></td><td>${st.warehouse}</td><td>${st.po_number}</td><td>${st.quantity}</td><td>${st.segment_km}</td><td>${segHrs}h${restInfo}</td><td class="text-danger fw-bold">${addDayName(st.latest_departure||'')}</td><td>${addDayName(st.arrival_time)}</td><td>${addDayName(st.adjusted_arrival)}</td><td>${(function(){
          try{var a=st.arrival_time.split(' '),b=st.adjusted_arrival.split(' ');
          var d1=new Date(a[0]+'T'+a[1]),d2=new Date(b[0]+'T'+b[1]);
          var wh=Math.round((d2-d1)/3600000*10)/10;
          return wh>0?wh+'h':'-';}catch(e){return '-';}
        })()}</td><td>${st.due_date}</td><td>${opHrs}</td></tr>`;
      });
      html += '</tbody></table>';
    });

    // Cost breakdown
    if (gr.routes.length > 0) {
      html += `
      <h6 class="mt-2"><i class="bi bi-cash-stack"></i> Cost Breakdown (비용 내역)</h6>
      <table class="table table-sm table-bordered small mb-2" style="max-width:600px;">
        <thead class="table-light"><tr><th>Route</th><th>Base ($)</th><th>Distance (km)</th><th>Dist Cost ($)</th><th>Total ($)</th></tr></thead>
        <tbody>`;
      gr.routes.forEach((rt, ri) => {
        var dc = (rt.cost - a.base_cost).toFixed(2);
        html += `<tr><td>${ri+1}</td><td>${a.base_cost.toLocaleString()}</td><td>${rt.distance_km.toLocaleString()}</td><td>${parseFloat(dc).toLocaleString()}</td><td class="fw-bold">${rt.cost.toLocaleString()}</td></tr>`;
      });
      html += `<tr class="table-dark"><td colspan="4" class="text-end fw-bold">Group Total</td><td class="fw-bold text-success">$${gr.total_cost.toLocaleString()}</td></tr>`;
      html += '</tbody></table>';
    }

    // Infeasible
    if (gr.infeasible && gr.infeasible.length > 0) {
      html += '<h6 class="mt-3 text-danger"><i class="bi bi-exclamation-triangle"></i> Infeasible POs</h6><ul class="small">';
      gr.infeasible.forEach(inf => {
        html += `<li><strong>${inf.po.po_number}</strong> (${inf.po.warehouse}, Qty: ${inf.po.quantity})`;
        if (inf.reasons && inf.reasons.length) {
          html += '<ul>' + inf.reasons.map(r => `<li class="text-danger">${r}</li>`).join('') + '</ul>';
        }
        html += '</li>';
      });
      html += '</ul>';
    }

    html += '</div>';
  });

  // Close Route View tab
  html += '</div><!-- end tab-routes -->';

  // === Load Plan Tab ===
  html += '<div class="tab-pane fade" id="tab-loadplan" role="tabpanel">';
  html += '<div class="card p-3 mb-3">';
  html += '<h6><i class="bi bi-table"></i> Load Plan (적재 계획)</h6>';
  html += '<div class="table-responsive">';
  html += '<table class="table table-sm table-bordered table-striped small mb-0" id="loadplan-table">';
  html += '<thead class="table-dark"><tr>';
  html += '<th>Route</th><th>PU from</th><th>SHIP TO</th><th>DC</th><th>PO #</th><th>Item #</th>';
  html += '<th>QTY<br>(PALLETS)</th><th>Loading</th><th>PICK UP</th><th>APPT (Delivery)</th>';
  html += '<th>Hours</th><th>Due Date</th><th>WAVE</th><th>CARRIER</th>';
  html += '</tr></thead><tbody>';

  // Build warehouse address lookup
  var whAddr = {};
  WAREHOUSES.forEach(w => { whAddr[w.name] = w.address || ''; });

  // Flatten all routes into load plan rows
  var allGroups = [result.ms_result, result.l1_result];
  var routeNum = 0;
  allGroups.forEach((gr, gi) => {
    if (!gr || !gr.routes || gr.routes.length === 0) return;
    var puFrom = (gr.group === 'MS-WH') ? 'MS' : 'L1';
    gr.routes.forEach((rt, ri) => {
      routeNum++;
      var stopCount = rt.schedule.length;
      rt.schedule.forEach((st, si) => {
        var addr = whAddr[st.warehouse] || st.warehouse;
        var segHrs = st.segment_hours || (st.segment_km / 80).toFixed(1);
        var pickupStr = (si === 0) ? rt.departure : '';
        var routeLabel = (si === 0) ? 'R' + routeNum : '';
        var puLabel = (si === 0) ? puFrom : '';
        var apptStr = st.adjusted_arrival || '';
        var badgeCls = (puFrom==='MS') ? 'bg-danger' : 'bg-primary';

        // Loading position: 1 stop='-', 2 stops=TAIL/NOSE, 3 stops=TAIL/MIDDLE/NOSE
        var loading = '-';
        if (stopCount === 2) {
          loading = (si === 0) ? 'TAIL' : 'NOSE';
        } else if (stopCount >= 3) {
          if (si === 0) loading = 'TAIL';
          else if (si === stopCount - 1) loading = 'NOSE';
          else loading = 'MIDDLE';
        }
        var loadOpts = ['', 'TAIL', 'MIDDLE', 'NOSE'];
        var loadSel = '';
        loadOpts.forEach(function(opt) {
          var lbl = opt || '-';
          var sel = (opt === loading || (opt === '' && loading === '-')) ? ' selected' : '';
          loadSel += '<option value="' + opt + '"' + sel + '>' + lbl + '</option>';
        });

        html += '<tr>';
        html += '<td class="fw-bold">' + routeLabel + '</td>';
        html += '<td><span class="badge ' + badgeCls + '">' + puLabel + '</span></td>';
        html += '<td style="white-space:pre-line;min-width:200px;font-size:0.8rem;">' + addr + '</td>';
        html += '<td><span class="badge bg-secondary">' + (st.dc_code||'') + '</span></td>';
        html += '<td>' + st.po_number + '</td>';
        html += '<td><input type="text" class="form-control form-control-sm" style="width:70px;" value="DSKT"></td>';
        html += '<td class="text-center fw-bold">' + st.quantity + '</td>';
        html += '<td><select class="form-select form-select-sm" style="width:80px;">' + loadSel + '</select></td>';
        html += '<td style="white-space:nowrap;">' + pickupStr + '</td>';
        html += '<td style="white-space:nowrap;">' + apptStr + '</td>';
        html += '<td class="text-center">' + segHrs + '</td>';
        html += '<td>' + st.due_date + '</td>';
        html += '<td><input type="number" class="form-control form-control-sm" style="width:50px;" value="2" min="1"></td>';
        html += '<td><input type="text" class="form-control form-control-sm" style="width:90px;" value=""></td>';
        html += '</tr>';
      });
    });
  });

  html += '</tbody></table></div></div></div><!-- end tab-loadplan -->';
  html += '</div><!-- end tab-content -->';

  // Update summary counts to match current state
  // Recalculate live totals from current routes
  var liveTotal = 0;
  var liveRoutes = 0;
  [result.ms_result, result.l1_result].forEach(gr => {
    if (gr && gr.routes) {
      gr.total_cost = gr.routes.reduce((s, r) => s + r.cost, 0);
      gr.total_routes = gr.routes.length;
      liveTotal += gr.total_cost;
      liveRoutes += gr.total_routes;
    }
  });

  // Save original cost on first render for comparison
  if (!window._originalTotal) {
    window._originalTotal = s.grand_total;
    window._originalRoutes = s.total_routes;
  }
  var origTotal = window._originalTotal;
  var diff = liveTotal - origTotal;
  var diffStr = '';
  if (Math.abs(diff) > 1) {
    var diffCls = diff > 0 ? 'text-danger' : 'text-success';
    var sign = diff > 0 ? '+' : '';
    diffStr = '<div class="small ' + diffCls + ' fw-bold">' + sign + '$' + Math.round(diff).toLocaleString() + ' vs original</div>';
  }
  var routeDiff = liveRoutes - window._originalRoutes;
  var routeDiffStr = '';
  if (routeDiff !== 0) {
    var rdCls = routeDiff > 0 ? 'text-danger' : 'text-success';
    var rdSign = routeDiff > 0 ? '+' : '';
    routeDiffStr = '<div class="small ' + rdCls + '">' + rdSign + routeDiff + ' vs original</div>';
  }

  // Grand summary (outside tabs)
  html += '<div class="card p-4 mb-4 border-start border-4 border-success">';
  html += '<div class="row text-center">';
  html += '<div class="col-md-3"><div class="fs-3 fw-bold">' + s.total_pos + '</div><div class="text-muted small">Total POs (총 주문)</div></div>';
  html += '<div class="col-md-3"><div class="fs-3 fw-bold">' + liveRoutes + '</div><div class="text-muted small">Total Routes (총 경로)</div>' + routeDiffStr + '</div>';
  html += '<div class="col-md-3"><div class="fs-3 fw-bold text-success">$' + Math.round(liveTotal).toLocaleString() + '</div><div class="text-muted small">Total Cost (총 비용)</div>' + diffStr + '</div>';
  html += '<div class="col-md-3"><div class="fs-3 fw-bold ' + (s.total_infeasible>0?'text-danger':'text-success') + '">' + s.total_infeasible + '</div><div class="text-muted small">Infeasible POs</div></div>';
  html += '</div>';
  html += '<div class="mt-3 small text-muted text-center">';
  html += s.total_infeasible === 0
    ? 'All POs can be fulfilled with the routes above. (모든 주문이 위 경로로 처리 가능합니다.)'
    : s.total_infeasible + ' PO(s) could not be scheduled. Consider adjusting due dates or receiving windows.';
  html += '</div></div>';

  document.getElementById('results-section').innerHTML = html;
  document.getElementById('results-section').style.display = 'block';
  document.getElementById('input-section').style.display = 'none';

  // Store result for export
  window._lastResult = result;
  window._lastPayload = payload;

  // Render map (reset instance since HTML was replaced)
  window._mapInstance = null; window._routeLayerGroup = null;
  renderMap(result);
}

function updateMapCosts(result) {
  var l1El = document.getElementById('map-cost-l1');
  var msEl = document.getElementById('map-cost-ms');
  var totalEl = document.getElementById('map-cost-total');
  if (!l1El || !msEl || !totalEl || !result) return;

  var l1 = result.l1_result;
  var ms = result.ms_result;

  var l1Cost = (l1 && l1.routes) ? l1.routes.reduce(function(s,r){return s+r.cost;},0) : 0;
  var msCost = (ms && ms.routes) ? ms.routes.reduce(function(s,r){return s+r.cost;},0) : 0;
  var totalCost = l1Cost + msCost;

  // Save initial costs (first render)
  if (!window._initL1Cost && (l1Cost > 0 || msCost > 0)) {
    window._initL1Cost = l1Cost;
    window._initMSCost = msCost;
    window._initTotalCost = totalCost;
  }
  if (!window._prevL1Cost) window._prevL1Cost = l1Cost;
  if (!window._prevMSCost) window._prevMSCost = msCost;

  var l1Prev = l1Cost - window._prevL1Cost;
  var msPrev = msCost - window._prevMSCost;
  var l1Init = l1Cost - (window._initL1Cost || l1Cost);
  var msInit = msCost - (window._initMSCost || msCost);
  var totalInit = totalCost - (window._initTotalCost || totalCost);
  var totalPrev = (l1Cost + msCost) - (window._prevL1Cost + window._prevMSCost);

  function diffTag(d, label) {
    if (Math.abs(d) < 1) return "";
    var sign = d > 0 ? "+" : "";
    var cls = d > 0 ? "color:#faa" : "color:#7f7";
    return " <span style=\"font-size:10px;" + cls + "\">" + label + sign + "$" + Math.round(d).toLocaleString() + "</span>";
  }

  if (l1Cost > 0) {
    l1El.style.display = "block";
    l1El.innerHTML = "L1: $" + Math.round(l1Cost).toLocaleString() + diffTag(l1Prev, "") + diffTag(l1Init, "초기");
  } else { l1El.style.display = "none"; }

  if (msCost > 0) {
    msEl.style.display = "block";
    msEl.innerHTML = "MS: $" + Math.round(msCost).toLocaleString() + diffTag(msPrev, "") + diffTag(msInit, "초기");
  } else { msEl.style.display = "none"; }

  totalEl.style.display = "block";
  totalEl.innerHTML = "Total: $" + Math.round(totalCost).toLocaleString() + diffTag(totalPrev, "") + diffTag(totalInit, "초기");

  window._prevL1Cost = l1Cost;
  window._prevMSCost = msCost;
}

function renderMap(result) {
  if (!window._mapInstance) {
    initMap();
  }
  drawRoutes(result);
}

function initMap() {
  var mapEl = document.getElementById('route-map');
  if (!mapEl) return;

  // Destroy existing map instance if any
  if (mapEl._leaflet_id) {
    mapEl._leaflet = null;
    mapEl.innerHTML = '';
    mapEl.removeAttribute('class');
    delete mapEl._leaflet_id;
  }

  // Restore correct size
  if (window._isFullscreen) {
    mapEl.style.height = '100vh';
    mapEl.style.borderRadius = '0';
  } else {
    mapEl.style.height = '75vh';
    mapEl.style.borderRadius = '8px';
  }

  var map = L.map('route-map', {
    maxBounds: [[24, -130], [50, -65]],
    maxBoundsViscosity: 1.0,
    minZoom: 4
  }).setView([39.8, -98.5], 4);
  window._mapInstance = map;
  window._routeLayerGroup = L.layerGroup().addTo(map);

  L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png', {
    attribution: '&copy; OpenStreetMap',
    maxZoom: 18
  }).addTo(map);
}

function drawRoutes(result) {
  var map = window._mapInstance;
  if (!map || !result || !result.coords) return;

  // Clear only route layers, keep tile layer
  window._routeLayerGroup.clearLayers();
  window._mapMarkers = [];
  window._mapSelected = new Set(); window._mapSelectOrder = [];
  window._selectedSeg = null;

  var coords = result.coords;
  var bounds = [];
  var routeColors = ['#e74c3c','#3498db','#2ecc71','#f39c12','#9b59b6','#1abc9c','#e67e22','#34495e'];
  var lg = window._routeLayerGroup;

  var groups = [result.ms_result, result.l1_result];
  var colorIdx = 0;

  for (var g = 0; g < groups.length; g++) {
    var gr = groups[g];
    if (!gr) continue;
    if (!gr.routes) gr.routes = [];

    var isMS = gr.group === 'MS-WH';
    var originKey = isMS ? '__ms_origin__' : '__l1_origin__';
    var originCoord = coords[originKey];
    if (!originCoord) continue;

    // Origin marker
    var oMIdx = window._mapMarkers.length;
    var originIcon = L.divIcon({
      html: '<div class="map-stop-marker" id="msm-' + oMIdx + '" style="background:' + (isMS?'#e74c3c':'#3498db') + ';color:#fff;border-radius:50%;width:30px;height:30px;display:flex;align-items:center;justify-content:center;font-weight:bold;font-size:14px;border:2px solid #fff;box-shadow:0 2px 5px rgba(0,0,0,0.3);cursor:pointer;">&#9733;</div>',
      className: '',
      iconSize: [30, 30],
      iconAnchor: [15, 15]
    });
    var originMarker = L.marker([originCoord.lat, originCoord.lon], {icon: originIcon})
      .bindPopup('<b>' + gr.group + ' Origin</b><br>' + (originCoord.label || ''));
    lg.addLayer(originMarker);
    bounds.push([originCoord.lat, originCoord.lon]);

    window._mapMarkers.push({
      marker: originMarker, gi: g, ri: -99, si: -1, color: isMS?'#e74c3c':'#3498db',
      po: '', wh: '', qty: 0, due: '', dc: '', isOrigin: true, groupName: gr.group
    });
    (function(idx) {
      originMarker.on('click', function(e) {
        L.DomEvent.stopPropagation(e);
        toggleMapSelect(idx);
      });
    })(oMIdx);

    for (var r = 0; r < gr.routes.length; r++) {
      var rt = gr.routes[r];
      var color = routeColors[colorIdx % routeColors.length];
      colorIdx++;

      var latlngs = [[originCoord.lat, originCoord.lon]];

      for (var si = 0; si < rt.schedule.length; si++) {
        var stop = rt.schedule[si];
        var wc = coords[stop.warehouse];
        if (!wc) continue;

        latlngs.push([wc.lat, wc.lon]);
        bounds.push([wc.lat, wc.lon]);

        var mIdx = window._mapMarkers.length;
        var stopIcon = L.divIcon({
          html: '<div class="map-stop-marker" id="msm-' + mIdx + '" style="background:' + color + ';color:#fff;border-radius:8px;min-width:28px;height:22px;padding:0 4px;display:flex;align-items:center;justify-content:center;font-weight:bold;font-size:10px;border:2px solid #fff;box-shadow:0 2px 4px rgba(0,0,0,0.3);cursor:pointer;white-space:nowrap;">' + (stop.dc_code || (si+1)) + '</div>',
          className: '',
          iconSize: [24, 24],
          iconAnchor: [12, 12]
        });
        var marker = L.marker([wc.lat, wc.lon], {icon: stopIcon})
          .bindPopup('<b>' + rt.schedule.map(function(s){return s.dc_code;}).join('\u2192') + ' / Stop ' + (si+1) + '</b><br>' + (stop.dc_code ? '<b>[' + stop.dc_code + ']</b> ' : '') + stop.warehouse + '<br>PO: ' + stop.po_number + '<br>Qty: ' + stop.quantity + '<br>Arrival: ' + stop.adjusted_arrival);
        lg.addLayer(marker);

        window._mapMarkers.push({
          marker: marker, gi: g, ri: r, si: si, color: color,
          po: stop.po_number, wh: stop.warehouse, qty: stop.quantity,
          due: stop.due_date, dc: stop.dc_code || ''
        });

        (function(idx) {
          marker.on('click', function(e) {
            L.DomEvent.stopPropagation(e);
            toggleMapSelect(idx);
          });
        })(mIdx);
      }

      // Route segments
      for (var seg = 0; seg < latlngs.length - 1; seg++) {
        var segLine = L.polyline([latlngs[seg], latlngs[seg+1]], {color: color, weight: 4, opacity: 0.8});
        segLine._routeTag = {gi: g, ri: r};
        lg.addLayer(segLine);
        (function(line, clr, gIdx, rIdx, segIdx) {
          line.on('mouseover', function() { this.setStyle({weight: 7, color: '#ff0', opacity: 1}); });
          line.on('mouseout', function() {
            if (window._selectedSeg && window._selectedSeg.line === line) {
              this.setStyle({weight: 6, color: '#f00', opacity: 1, dashArray: '10,6'});
            } else {
              this.setStyle({weight: 4, color: clr, opacity: 0.8, dashArray: null});
            }
          });
          line.on('click', function(e) {
            L.DomEvent.stopPropagation(e);
            if (window._selectedSeg && window._selectedSeg.line !== line) {
              window._selectedSeg.line.setStyle({weight: 4, color: window._selectedSeg.clr, opacity: 0.8, dashArray: null});
            }
            if (window._selectedSeg && window._selectedSeg.line === line) {
              line.setStyle({weight: 4, color: clr, opacity: 0.8, dashArray: null});
              window._selectedSeg = null;
            } else {
              line.setStyle({weight: 6, color: '#f00', opacity: 1, dashArray: '10,6'});
              window._selectedSeg = {line: line, clr: clr, gi: gIdx, ri: rIdx, seg: segIdx};
            }
            updateMapActionBar();
          });
        })(segLine, color, g, r, seg);
      }

      // Route label
      var labelLat, labelLon;
      if (isMS) { labelLat = 44 - r * 1.8; labelLon = -62; }
      else { labelLat = 44 - r * 1.8; labelLon = -128; }
      var routeLabel = L.marker([labelLat, labelLon], {
        icon: L.divIcon({
          html: '<div style="background:#fff;color:' + color + ';border:2px solid ' + color + ';border-radius:6px;min-width:80px;padding:0 8px;height:24px;line-height:24px;font-size:11px;font-weight:bold;text-align:center;cursor:pointer;white-space:nowrap;">' + rt.schedule.map(function(s){return s.dc_code||'';}).join('\u2192') + ' (' + rt.total_units + ')' + '</div>',
          className: '',
          iconSize: [120, 24],
          iconAnchor: [50, 12]
        })
      });
      lg.addLayer(routeLabel);
      (function(gIdx, rIdx, clr) {
        routeLabel.on('click', function() { showRouteModal(gIdx, rIdx, clr); });
      })(g, r, color);
    }
  }

  // Detached routes
  if (result._detached && result._detached.length > 0) {
    result._detached.forEach(function(dt, di) {
      var detLatlngs = [];
      dt.schedule.forEach(function(stop, si) {
        var wc = coords[stop.warehouse];
        if (!wc) return;
        bounds.push([wc.lat, wc.lon]);
        detLatlngs.push([wc.lat, wc.lon]);
        var mIdx = window._mapMarkers.length;
        var detIcon = L.divIcon({
          html: '<div class="map-stop-marker" id="msm-' + mIdx + '" style="background:#888;color:#fff;border-radius:8px;min-width:28px;height:22px;padding:0 4px;display:flex;align-items:center;justify-content:center;font-weight:bold;font-size:10px;border:3px dashed #f39c12;box-shadow:0 2px 4px rgba(0,0,0,0.3);cursor:pointer;white-space:nowrap;">' + (stop.dc_code || '?') + '</div>',
          className: '',
          iconSize: [28, 22],
          iconAnchor: [14, 11]
        });
        var marker = L.marker([wc.lat, wc.lon], {icon: detIcon})
          .bindPopup('<b>Detached</b><br>' + (stop.dc_code ? '<b>[' + stop.dc_code + ']</b> ' : '') + stop.warehouse + '<br>PO: ' + stop.po_number + '<br>Qty: ' + stop.quantity);
        lg.addLayer(marker);

        window._mapMarkers.push({
          marker: marker, gi: -1, ri: di, si: si, color: '#888',
          po: stop.po_number, wh: stop.warehouse, qty: stop.quantity,
          due: stop.due_date, dc: stop.dc_code || '', detached: true
        });

        (function(idx) {
          marker.on('click', function(e) {
            L.DomEvent.stopPropagation(e);
            toggleMapSelect(idx);
          });
        })(mIdx);
      });
      if (detLatlngs.length > 1) {
        var detLine = L.polyline(detLatlngs, {color: '#f39c12', weight: 3, opacity: 0.8, dashArray: '8,8'});
        detLine._routeTag = {gi: -1, ri: di};
        lg.addLayer(detLine);
      }
    });
  }

  if (bounds.length > 0) {
    map.fitBounds(bounds, {padding: [30, 30]});
  }
  updateMapActionBar();
  updateMapCosts(result);
}


function fetchTimeout(url, options, timeout) {
  timeout = timeout || 60000;
  return Promise.race([
    fetch(url, options),
    new Promise(function(_, reject) {
      setTimeout(function() { reject(new Error('요청 시간 초과 (' + (timeout/1000) + '초)')); }, timeout);
    })
  ]).then(function(resp) {
    var ct = resp.headers.get('content-type') || '';
    if (!ct.includes('application/json')) {
      throw new Error('서버 응답 오류 (nginx timeout 가능성). 직접 접속: http://10.0.30.12:8105');
    }
    return resp;
  });
}

function refreshAfterEdit(result) {
  // Clear stale segment selection (lines are recreated)
  window._selectedSeg = null;
  // Update map layers only (no full HTML rebuild)
  drawRoutes(result);
  // Update cost displays
  updateMapCosts(result);
  // Auto-run schedule after edit
  window._autoScheduleFromEdit = true;
  setTimeout(function() { autoSchedule(); }, 300);
  // Update summary totals in the DOM
  var liveTotal = 0, liveRoutes = 0;
  [result.ms_result, result.l1_result].forEach(function(gr) {
    if (gr && gr.routes) {
      gr.total_cost = gr.routes.reduce(function(s,r){return s+r.cost;}, 0);
      gr.total_routes = gr.routes.length;
      liveTotal += gr.total_cost;
      liveRoutes += gr.total_routes;
    }
  });
}

function toggleMapSelect(idx) {
  if (!window._mapSelectOrder) window._mapSelectOrder = [];
  if (window._mapSelected.has(idx)) {
    // Deselect
    window._mapSelected.delete(idx);
    window._mapSelectOrder = window._mapSelectOrder.filter(i => i !== idx);
  } else if (window._mapSelected.size >= 2) {
    // 3rd click: remove oldest, keep recent + new one
    var oldest = window._mapSelectOrder.shift();
    window._mapSelected.delete(oldest);
    var oldEl = document.getElementById('msm-' + oldest);
    if (oldEl) { oldEl.style.border = '2px solid #fff'; oldEl.style.boxShadow = '0 2px 4px rgba(0,0,0,0.3)'; oldEl.style.transform = 'scale(1)'; }
    window._mapSelected.add(idx);
    window._mapSelectOrder.push(idx);
  } else {
    window._mapSelected.add(idx);
    window._mapSelectOrder.push(idx);
  }
  // Update marker visual
  var el = document.getElementById('msm-' + idx);
  if (el) {
    if (window._mapSelected.has(idx)) {
      el.style.border = '3px solid #ff0';
      el.style.boxShadow = '0 0 10px 3px rgba(255,255,0,0.7)';
      el.style.transform = 'scale(1.3)';
    } else {
      el.style.border = '2px solid #fff';
      el.style.boxShadow = '0 2px 4px rgba(0,0,0,0.3)';
      el.style.transform = 'scale(1)';
    }
  }
  updateMapActionBar();
}

function updateMapActionBar() {
  var bar = document.getElementById('map-action-bar');
  var topBadge = document.getElementById('map-sel-count-top');
  var segBadge = document.getElementById('seg-selected');
  var cnt = window._mapSelected ? window._mapSelected.size : 0;
  var hasSeg = !!window._selectedSeg;
  if (cnt > 0 || hasSeg) {
    bar.style.display = 'flex';
    document.getElementById('map-sel-count').textContent = cnt;
    if (topBadge) { topBadge.style.display = cnt > 0 ? 'inline' : 'none'; topBadge.textContent = cnt + ' selected'; }
    if (segBadge) segBadge.style.display = hasSeg ? 'inline' : 'none';
  } else {
    bar.style.display = 'none';
    if (topBadge) topBadge.style.display = 'none';
    if (segBadge) segBadge.style.display = 'none';
  }
}

function mapClearSelection() {
  if (window._mapSelected) {
    window._mapSelected.forEach(idx => {
      var el = document.getElementById('msm-' + idx);
      if (el) { el.style.border = '2px solid #fff'; el.style.boxShadow = '0 2px 4px rgba(0,0,0,0.3)'; el.style.transform = 'scale(1)'; }
    });
  }
  window._mapSelected = new Set(); window._mapSelectOrder = [];
  window._mapMarkers = window._mapMarkers || [];
  // Clear selected segment
  if (window._selectedSeg) {
    window._selectedSeg.line.setStyle({weight: 4, color: window._selectedSeg.clr, opacity: 0.8, dashArray: null});
    window._selectedSeg = null;
  }
  updateMapActionBar();
}

async function mapCutSegment(gi, ri, segIdx) {
  // Cut route at segment segIdx: split into two routes
  // segIdx=0 means cut between origin and stop1 → [stop1...N] becomes separate
  // segIdx=1 means cut between stop1 and stop2 → [stop1] stays, [stop2...N] becomes separate
  var result = window._lastResult;
  if (!result) return;

  var gr = [result.ms_result, result.l1_result][gi];
  if (!gr || !gr.routes[ri]) return;
  var rt = gr.routes[ri];

  // segIdx 0 = origin→stop0, 1 = stop0→stop1, etc.
  // Cut after segIdx: part1 = stops[0..segIdx-1], part2 = stops[segIdx..end]
  // If segIdx=0: part1 is empty (origin→stop0 cut), all stops become part2
  var sched = rt.schedule;

  // segIdx=0 and 1 stop: cut origin→stop = remove route from this group
  if (segIdx === 0 && sched.length === 1) {
    // Proceed with cut
    // Remove from group
    var detached = gr.routes.splice(ri, 1)[0];
    gr.total_cost = gr.routes.reduce((s, r) => s + r.cost, 0);
    gr.total_routes = gr.routes.length;
    // Add to "unassigned" list for display
    if (!result._detached) result._detached = [];
    result._detached.push(detached);
    window._mapSelected = new Set(); window._mapSelectOrder = [];
    refreshAfterEdit(result);
    return;
  }

  var part1 = sched.slice(0, segIdx);   // stops before the cut
  var part2 = sched.slice(segIdx);      // stops after the cut

  if (part1.length === 0) {
    // Cut origin→stop0: detach entire route from this group (MN→NT stays connected)
    // Proceed with cut
    var detached = gr.routes.splice(ri, 1)[0];
    gr.total_cost = gr.routes.reduce((s, r) => s + r.cost, 0);
    gr.total_routes = gr.routes.length;
    if (!result._detached) result._detached = [];
    result._detached.push(detached);
    window._mapSelected = new Set(); window._mapSelectOrder = [];
    if (result._detached && result._detached.length === 0) delete result._detached;
    refreshAfterEdit(result);
    return;
  } else if (part2.length === 0) {
    showError('이 위치에서는 분리할 수 없습니다.'); return;
  }

  try {
    // Recalc part1 as the original route
    var resp1 = await fetchTimeout('/api/recalc-route', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({
        po_numbers: part1.map(s => s.po_number),
        warehouses: part1.map(s => s.warehouse),
        quantities: part1.map(s => s.quantity),
        due_dates: part1.map(s => s.due_date),
        pickup_time: rt.departure, group: gr.group,
      }),
    });
    var d1 = await resp1.json();
    if (d1.success) {
      gr.routes[ri].schedule = d1.route.schedule;
      gr.routes[ri].distance_km = d1.route.distance_km;
      gr.routes[ri].cost = d1.route.cost;
      gr.routes[ri].total_units = part1.reduce((s,st) => s + st.quantity, 0);
      gr.routes[ri].po_numbers = part1.map(s => s.po_number);
      gr.routes[ri].warehouses = part1.map(s => s.warehouse);
    }

    // Part2 becomes detached (no origin line, shown as dashed)
    var detRoute = {
      schedule: part2,
      po_numbers: part2.map(s => s.po_number),
      warehouses: part2.map(s => s.warehouse),
      departure: rt.departure,
      total_units: part2.reduce((s,st) => s + st.quantity, 0),
      distance_km: 0,
      cost: 0,
    };
    if (!result._detached) result._detached = [];
    result._detached.push(detRoute);

    // Recalc totals
    gr.total_cost = gr.routes.reduce((s, r) => s + r.cost, 0);
    gr.total_routes = gr.routes.length;

    window._mapSelected = new Set(); window._mapSelectOrder = [];
    // Clean up empty detached array
    if (result._detached && result._detached.length === 0) delete result._detached;
    refreshAfterEdit(result);
  } catch(e) { showError('Split error: ' + e.message); }
}

async function mapMerge() {
  // Show loading
  var loadEl = document.createElement('div');
  loadEl.id = 'merge-loading';
  loadEl.style.cssText = 'position:fixed;top:50%;left:50%;transform:translate(-50%,-50%);z-index:20000;background:rgba(0,0,0,0.8);color:#fff;padding:16px 24px;border-radius:12px;font-size:14px;';
  loadEl.innerHTML = '<i class="bi bi-hourglass-split"></i> 처리 중...';
  document.body.appendChild(loadEl);
  try { await _doMapMerge(); } finally { var el = document.getElementById('merge-loading'); if (el) el.remove(); }
}
async function _doMapMerge() {
  // Merge: combine entire routes that contain selected markers
  var result = window._lastResult;
  if (!result || !window._mapSelected || window._mapSelected.size === 0) {
    showError('합칠 루트의 마커를 선택하세요.'); return;
  }

  var grList = [result.ms_result, result.l1_result]; // unfiltered - gi matches renderMap index
  var selected = [];
  window._mapSelected.forEach(idx => {
    var m = window._mapMarkers[idx];
    if (!m) { console.error('mapMerge: marker idx=' + idx + ' not found in _mapMarkers (length=' + window._mapMarkers.length + ')'); return; }
    selected.push(m);
  });
  console.log('mapMerge: selected=' + selected.length, selected.map(s => 'gi='+s.gi+' ri='+s.ri+' dc='+s.dc+' origin='+!!s.isOrigin+' detached='+!!s.detached));
  if (selected.length === 0) { showError('선택된 마커가 없습니다.'); return; }

  // Check same group
  var gis = new Set(selected.map(s => s.gi));
  // Allow cross-group merge

  // Separate detached, origin, and grouped markers
  var detachedPOs = [];
  var originMarkers = [];
  var groupedSelected = [];
  selected.forEach(s => {
    if (s.detached) {
      detachedPOs.push(s);
    } else if (s.isOrigin) {
      originMarkers.push(s);
    } else {
      groupedSelected.push(s);
    }
  });

  // Special case: all detached (no origin, no grouped) = merge into one detached route
  if (detachedPOs.length >= 2 && originMarkers.length === 0 && groupedSelected.length === 0) {
    // Combine all detached POs into one detached route
    var mergedSchedule = [];
    var usedDetRis = new Set();
    detachedPOs.forEach(dp => {
      usedDetRis.add(dp.ri);
      if (result._detached && result._detached[dp.ri]) {
        var drt = result._detached[dp.ri];
        // Find specific stop
        var st = drt.schedule[dp.si];
        if (st) mergedSchedule.push(st);
      }
    });
    if (mergedSchedule.length > 0) {
      var totalUnits = mergedSchedule.reduce((s,st) => s + st.quantity, 0);
      if (totalUnits > 60) {
        showError('합산 ' + totalUnits + ' units — 최대 용량(60) 초과!');
        return;
      }
      // Remove used detached routes
      var dris = [...usedDetRis].sort((a,b) => b-a);
      dris.forEach(ri => { if (result._detached) result._detached.splice(ri, 1); });
      // Add merged detached
      var newDet = {
        schedule: mergedSchedule,
        po_numbers: mergedSchedule.map(s => s.po_number),
        warehouses: mergedSchedule.map(s => s.warehouse),
        departure: '',
        total_units: mergedSchedule.reduce((s,st) => s + st.quantity, 0),
        distance_km: 0, cost: 0,
      };
      if (!result._detached) result._detached = [];
      result._detached.push(newDet);
      if (result._detached.length === 0) delete result._detached;
      window._mapSelected = new Set(); window._mapSelectOrder = [];
      refreshAfterEdit(result);
    }
    return;
  }

  // Special case: detached + grouped (no origin) = add detached POs INTO the grouped route
  var nonOriginGrouped = groupedSelected.filter(s => s.ri >= 0);
  console.log('mapMerge paths: detached='+detachedPOs.length+' origin='+originMarkers.length+' grouped='+groupedSelected.length+' nonOriginGrouped='+nonOriginGrouped.length);
  if (detachedPOs.length > 0 && nonOriginGrouped.length > 0 && originMarkers.length === 0) {
    // Get target: the grouped marker's route
    var tgt = nonOriginGrouped[0];
    var tgtGr = grList[tgt.gi];
    if (!tgtGr || !tgtGr.routes[tgt.ri]) { showError('Route not found'); return; }
    var tgtRt = tgtGr.routes[tgt.ri];

    // Collect all detached POs from their entire routes
    var detPOs = [];
    var usedDetRis = new Set();
    detachedPOs.forEach(dp => {
      usedDetRis.add(dp.ri);
      if (result._detached && result._detached[dp.ri]) {
        result._detached[dp.ri].schedule.forEach(st => {
          detPOs.push({ po: st.po_number, wh: st.warehouse, qty: st.quantity, due: st.due_date });
        });
      }
    });

    // Combine: existing route POs + detached POs
    var allMergePOs = [];
    tgtRt.schedule.forEach(st => {
      allMergePOs.push({ po: st.po_number, wh: st.warehouse, qty: st.quantity, due: st.due_date });
    });
    detPOs.forEach(p => allMergePOs.push(p));

    var totalUnits = allMergePOs.reduce((s,p) => s + p.qty, 0);
    console.log('detached+grouped merge: allMergePOs='+allMergePOs.length+' totalUnits='+totalUnits+' detPOs='+detPOs.length+' route POs='+tgtRt.schedule.length);
    if (totalUnits > 60) {
      showError('합산 ' + totalUnits + ' units — 최대 용량(60) 초과!');
      return;
    }

    try {
      var resp5 = await fetchTimeout('/api/recalc-route', {
        method: 'POST', headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({
          po_numbers: allMergePOs.map(p => p.po), warehouses: allMergePOs.map(p => p.wh),
          quantities: allMergePOs.map(p => p.qty), due_dates: allMergePOs.map(p => p.due),
          pickup_time: tgtRt.departure, group: tgtGr.group,
        }),
      });
      var d5 = await resp5.json();
      console.log('detached+grouped recalc result:', d5.success, d5.error||'ok');
      if (d5.success) {
        tgtRt.schedule = d5.route.schedule;
        tgtRt.distance_km = d5.route.distance_km;
        tgtRt.cost = d5.route.cost;
        tgtRt.total_units = allMergePOs.reduce((s,p) => s + p.qty, 0);
        tgtRt.po_numbers = allMergePOs.map(p => p.po);
        tgtRt.warehouses = allMergePOs.map(p => p.wh);
      }

      // Remove used detached routes
      var dris2 = [...usedDetRis].sort((a,b) => b-a);
      dris2.forEach(ri => { if (result._detached) result._detached.splice(ri, 1); });
      if (result._detached && result._detached.length === 0) delete result._detached;

      tgtGr.total_cost = tgtGr.routes.reduce((s,r) => s + r.cost, 0);
      tgtGr.total_routes = tgtGr.routes.length;
      window._mapSelected = new Set(); window._mapSelectOrder = [];
      refreshAfterEdit(result);
    } catch(e) { showError('Merge error: ' + e.message); }
    return;
  }

  // Special case: origin + detached = add detached POs to that group as new route
  if (originMarkers.length > 0 && detachedPOs.length > 0 && nonOriginGrouped.length === 0) {
    console.log('origin+detached merge path');
    var targetGi = originMarkers[0].gi;
    var targetGr = grList[targetGi];
    if (!targetGr) { showError('Group not found (gi='+targetGi+')'); return; }

    try {
      // Collect ALL POs from detached routes that contain any selected marker
      var allDetPOs = [];
      var affectedDetRis = [...new Set(detachedPOs.map(d => d.ri))];
      affectedDetRis.forEach(ri => {
        var drt = result._detached[ri];
        if (!drt) return;
        drt.schedule.forEach(st => {
          allDetPOs.push({ po: st.po_number, wh: st.warehouse, qty: st.quantity, due: st.due_date });
        });
      });

      // Check capacity
      var totalUnits = allDetPOs.reduce((s,p) => s + p.qty, 0);
      console.log('origin+detached: allDetPOs='+allDetPOs.length+' totalUnits='+totalUnits);
      if (totalUnits > 60) {
        showError('합산 ' + totalUnits + ' units — 최대 용량(60) 초과!');
        return;
      }

      // Create route with all selected detached POs
      console.log('origin+detached: calling recalc-route, group='+targetGr.group);
      var resp = await fetchTimeout('/api/recalc-route', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({
          po_numbers: allDetPOs.map(p => p.po),
          warehouses: allDetPOs.map(p => p.wh),
          quantities: allDetPOs.map(p => p.qty),
          due_dates: allDetPOs.map(p => p.due),
          pickup_time: new Date().toISOString().slice(0,10) + ' 10:00',
          group: targetGr.group,
        }),
      });
      var data = await resp.json();
      if (data.success) {
        var nr = data.route;
        nr.po_numbers = allDetPOs.map(p => p.po);
        nr.warehouses = allDetPOs.map(p => p.wh);
        nr.departure = new Date().toISOString().slice(0,10) + ' 10:00';
        nr.total_units = allDetPOs.reduce((s,p) => s + p.qty, 0);
        targetGr.routes.push(nr);
      }

      // Remove entire detached routes that were used
      if (result._detached) {
        var dris = [...new Set(detachedPOs.map(d => d.ri))].sort((a,b) => b-a);
        dris.forEach(ri => result._detached.splice(ri, 1));
        if (result._detached.length === 0) delete result._detached;
      }
      targetGr.total_cost = targetGr.routes.reduce((s, r) => s + r.cost, 0);
      targetGr.total_routes = targetGr.routes.length;
      window._mapSelected = new Set(); window._mapSelectOrder = [];
      refreshAfterEdit(result);
    } catch(e) { showError('Error: ' + e.message); }
    return;
  }

  // Collect affected routes per group (exclude origins ri=-99)
  var affectedByGroup = {}; // gi -> Set of ri
  groupedSelected.forEach(s => {
    if (s.ri < 0) return; // skip origins
    if (!affectedByGroup[s.gi]) affectedByGroup[s.gi] = new Set();
    affectedByGroup[s.gi].add(s.ri);
  });

  var totalAffected = 0;
  for (var k in affectedByGroup) totalAffected += affectedByGroup[k].size;
  totalAffected += detachedPOs.length;

  console.log('mapMerge: totalAffected='+totalAffected+' affectedByGroup='+JSON.stringify(Object.fromEntries(Object.entries(affectedByGroup).map(([k,v])=>[k,[...v]]))));
  if (totalAffected < 2) { showError('2개 이상의 마커를 선택하세요.'); return; }

  // Determine target group
  var uniqueGis = [...new Set(groupedSelected.filter(s => s.ri >= 0).map(s => s.gi))];
  var targetGi;
  if (uniqueGis.length > 1) {
    // Cross-group merge: ask user which origin to use
    var grNames = uniqueGis.map(gi => grList[gi] ? grList[gi].group : 'Group'+gi);
    var choice = prompt('어느 Origin에 연결할까요?\n' + grNames.map((n,i) => (i+1)+'. '+n).join('\n') + '\n\n번호 입력 (또는 0=detached):');
    if (choice === null) return;
    choice = parseInt(choice);
    if (choice === 0) {
      // Make it detached instead
      var mergedSchedule = [];
      groupedSelected.forEach(s => {
        if (s.ri < 0) return;
        var gr = grList[s.gi]; if (!gr || !gr.routes[s.ri]) return;
        var stop = gr.routes[s.ri].schedule[s.si];
        if (stop) mergedSchedule.push(stop);
      });
      detachedPOs.forEach(dp => {
        if (result._detached && result._detached[dp.ri]) {
          result._detached[dp.ri].schedule.forEach(st => mergedSchedule.push(st));
        }
      });
      // Remove from original routes
      var poSet = new Set(mergedSchedule.map(s => s.po_number));
      groupedSelected.forEach(s => {
        if (s.ri < 0) return;
        var gr = grList[s.gi]; if (!gr || !gr.routes[s.ri]) return;
        var rt = gr.routes[s.ri];
        rt.schedule = rt.schedule.filter(st => !poSet.has(st.po_number));
        if (rt.schedule.length === 0) { gr.routes.splice(s.ri, 1); }
        gr.total_cost = gr.routes.reduce((s,r) => s + r.cost, 0);
      });
      if (!result._detached) result._detached = [];
      result._detached.push({schedule: mergedSchedule, po_numbers: mergedSchedule.map(s=>s.po_number), warehouses: mergedSchedule.map(s=>s.warehouse), departure:'', total_units: mergedSchedule.reduce((s,st)=>s+st.quantity,0), distance_km:0, cost:0});
      window._mapSelected = new Set(); window._mapSelectOrder = [];
      refreshAfterEdit(result);
      return;
    }
    if (choice >= 1 && choice <= uniqueGis.length) {
      targetGi = uniqueGis[choice - 1];
    } else {
      targetGi = uniqueGis[0];
    }
  } else {
    targetGi = groupedSelected.length > 0 ? groupedSelected[0].gi : 0;
  }
  var targetGr = grList[targetGi];
  if (!targetGr) { targetGi = grList[0] ? 0 : 1; targetGr = grList[targetGi]; }
  if (!targetGr) { showError('Target group not found'); return; }

  // Collect ONLY the selected marker POs (not entire routes)
  var allPOs = [];
  var pickups = [];
  groupedSelected.forEach(s => {
    if (s.ri < 0) return;
    var gr = grList[s.gi];
    if (!gr || !gr.routes[s.ri]) return;
    var rt = gr.routes[s.ri];
    pickups.push(rt.departure);
    // Only add this specific stop's PO
    var stop = rt.schedule[s.si];
    if (stop) {
      allPOs.push({ po: stop.po_number, wh: stop.warehouse, qty: stop.quantity, due: stop.due_date });
    }
  });
  // Add detached POs
  detachedPOs.forEach(dp => {
    allPOs.push({ po: dp.po, wh: dp.wh, qty: dp.qty, due: dp.due });
    // Find pickup from detached route
    if (result._detached && result._detached[dp.ri]) {
      var drt = result._detached[dp.ri];
      pickups.push((drt && drt.departure) || (drt && drt.schedule && drt.schedule[0] && drt.schedule[0].due_date + ' 10:00') || new Date().toISOString().slice(0,10) + ' 10:00');
    }
  });

  var totalUnits = allPOs.reduce((s, p) => s + p.qty, 0);
  if (totalUnits > 60) {
    showError('합산 ' + totalUnits + ' units — 최대 용량(60) 초과!');
    return;
  }

  var pickup = pickups.length > 0 ? pickups.sort()[0] : new Date().toISOString().slice(0,10) + ' 10:00';
  console.log('mapMerge: allPOs=' + allPOs.length + ' pickup=' + pickup + ' group=' + targetGr.group);

  try {
    // Create merged route using target group's origin
    var resp = await fetchTimeout('/api/recalc-route', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({
        po_numbers: allPOs.map(p => p.po),
        warehouses: allPOs.map(p => p.wh),
        quantities: allPOs.map(p => p.qty),
        due_dates: allPOs.map(p => p.due),
        pickup_time: pickup, group: targetGr.group,
      }),
    });
    var data = await resp.json();
    if (!data.success) { showError('Merge failed: ' + (data.error || '')); return; }

    // Remove only the selected POs from their original routes (not entire routes)
    var selectedPOset = new Set(allPOs.map(p => p.po));
    for (var giKey in affectedByGroup) {
      var gr = grList[parseInt(giKey)];
      if (!gr) continue;
      var routesToDelete = [];
      var risArr = [...affectedByGroup[giKey]];
      for (var idx = 0; idx < risArr.length; idx++) {
        var ri = risArr[idx];
        var rt = gr.routes[ri];
        if (!rt) continue;
        var remaining = rt.schedule.filter(st => !selectedPOset.has(st.po_number));
        if (remaining.length === 0) {
          routesToDelete.push(ri);
        } else {
          var resp3 = await fetchTimeout('/api/recalc-route', {
            method: 'POST',
            headers: {'Content-Type': 'application/json'},
            body: JSON.stringify({
              po_numbers: remaining.map(s => s.po_number),
              warehouses: remaining.map(s => s.warehouse),
              quantities: remaining.map(s => s.quantity),
              due_dates: remaining.map(s => s.due_date),
              pickup_time: rt.departure, group: gr.group,
            }),
          });
          var d3 = await resp3.json();
          if (d3.success) {
            rt.schedule = d3.route.schedule;
            rt.distance_km = d3.route.distance_km;
            rt.cost = d3.route.cost;
            rt.total_units = remaining.reduce((s,st) => s + st.quantity, 0);
            rt.po_numbers = remaining.map(s => s.po_number);
            rt.warehouses = remaining.map(s => s.warehouse);
          }
        }
      }
      routesToDelete.sort((a,b) => b-a).forEach(ri => gr.routes.splice(ri, 1));
      gr.total_cost = gr.routes.reduce((s, r) => s + r.cost, 0);
      gr.total_routes = gr.routes.length;
    }
    // Remove used detached POs (only selected ones, keep rest in detached route)
    if (detachedPOs.length > 0 && result._detached) {
      var selectedDetPOs = new Set(detachedPOs.map(d => d.po));
      var affDetRis = [...new Set(detachedPOs.map(d => d.ri))].sort((a,b) => b-a);
      affDetRis.forEach(ri => {
        var drt = result._detached[ri];
        if (!drt) return;
        var remaining = drt.schedule.filter(st => !selectedDetPOs.has(st.po_number));
        if (remaining.length === 0) {
          result._detached.splice(ri, 1);
        } else {
          drt.schedule = remaining;
          drt.po_numbers = remaining.map(s => s.po_number);
          drt.warehouses = remaining.map(s => s.warehouse);
          drt.total_units = remaining.reduce((s,st) => s + st.quantity, 0);
        }
      });
      if (result._detached && result._detached.length === 0) delete result._detached;
    }

    // Add merged route to target group
    var merged = data.route;
    merged.po_numbers = allPOs.map(p => p.po);
    merged.warehouses = allPOs.map(p => p.wh);
    merged.departure = pickup;
    merged.total_units = totalUnits;
    targetGr.routes.push(merged);
    targetGr.total_cost = targetGr.routes.reduce((s, r) => s + r.cost, 0);
    targetGr.total_routes = targetGr.routes.length;

    window._mapSelected = new Set(); window._mapSelectOrder = [];
    refreshAfterEdit(result);
  } catch(e) { showError('Merge error: ' + e.message); }
}

function addDayName(dtStr) {
  // "2026-03-16 04:00 EDT" → "2026-03-16(Mon) 04:00 EDT"
  if (!dtStr || dtStr.length < 10) return dtStr;
  var datepart = dtStr.substring(0, 10);
  var rest = dtStr.substring(10);
  var d = new Date(datepart + 'T12:00:00');
  var days = ['Sun','Mon','Tue','Wed','Thu','Fri','Sat'];
  var day = days[d.getDay()];
  return datepart + '(' + day + ')' + rest;
}

async function autoSchedule() {
  var result = window._lastResult;
  if (!result) { alert('결과가 없습니다.'); return; }

  var btn = (typeof event !== 'undefined' && event && event.target) ? event.target.closest('button') : null;
  var origText = btn ? btn.innerHTML : '';
  if (btn) { btn.innerHTML = '<i class="bi bi-hourglass-split"></i> 계산 중...'; btn.disabled = true; }

  try {
    var allGroups = [result.ms_result, result.l1_result];
    for (var gi = 0; gi < allGroups.length; gi++) {
      var gr = allGroups[gi];
      if (!gr || !gr.routes || gr.routes.length === 0) continue;

      var resp = await fetch('/api/auto-schedule', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({routes: gr.routes, group: gr.group}),
      });
      var data = await resp.json();
      if (!data.success) continue;

      // Apply scheduled pickups and update routes
      data.schedules.forEach(function(sched, ri) {
        if (!sched.pickup || !gr.routes[ri]) return;
        var rt = gr.routes[ri];
        rt.departure = sched.pickup;
        // Update schedule from simulation
        if (sched.simulated_route && sched.simulated_route.schedule) {
          rt.schedule = sched.simulated_route.schedule;
          rt.distance_km = sched.simulated_route.distance_km;
          rt.cost = sched.simulated_route.cost;
        }
        rt._auto_schedule = {
          pickup: sched.pickup,
          pickup_day: sched.pickup_day,
          departure: sched.departure,
          departure_day: sched.departure_day,
          friday_rule: sched.friday_rule,
          travel_hours_first: sched.travel_hours_first_stop,
          target_arrival: sched.target_arrival,
        };
      });

      // Recalc totals
      gr.total_cost = gr.routes.reduce(function(s, r) { return s + r.cost; }, 0);
    }

    if (btn) { btn.innerHTML = origText; btn.disabled = false; }

    // Show summary
    var msg = '스케줄 자동결정 완료!\n\n';
    allGroups.forEach(function(gr) {
      if (!gr || !gr.routes || gr.routes.length === 0) return;
      msg += gr.group + ':\n';
      gr.routes.forEach(function(rt, ri) {
        var info = rt._auto_schedule || {};
        msg += '  R' + (ri+1) + ': 픽업 ' + (info.pickup||rt.departure) + '(' + (info.pickup_day||'') + ') → 출발 ' + (info.departure||rt.departure) + '(' + (info.departure_day||'') + ')';
        if (info.friday_rule) msg += ' [장거리]';
        msg += '\n';
      });
      msg += '\n';
    });
    console.log(msg);

    // If in fullscreen or called from refreshAfterEdit, just update map
    if (window._isFullscreen || window._autoScheduleFromEdit) {
      window._autoScheduleFromEdit = false;
      drawRoutes(result);
      updateMapCosts(result);
    } else {
      renderResults(result, window._lastPayload);
    }
  } catch(e) {
    if (btn) { btn.innerHTML = origText; btn.disabled = false; }
    alert('스케줄 결정 오류: ' + e.message);
  }
}

async function learnResult() {
  var result = window._lastResult;
  if (!result) { alert('결과가 없습니다.'); return; }

  var allGroups = [result.ms_result, result.l1_result];
  var learned = 0;
  for (var g = 0; g < allGroups.length; g++) {
    var gr = allGroups[g];
    if (!gr || !gr.routes || gr.routes.length === 0) continue;
    try {
      var resp = await fetch('/api/learn', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({routes: gr.routes, group: gr.group}),
      });
      var data = await resp.json();
      if (data.success) learned++;
    } catch(e) {}
  }
  if (learned > 0) {
    alert('학습 완료! 다음 최적화 시 이 조합 패턴이 반영됩니다.\n\n학습된 DC 페어링은 향후 최적화에서 보너스를 받습니다.');
  } else {
    alert('학습할 결과가 없습니다.');
  }
}

function toggleAIChat() {
  var panel = document.getElementById('ai-chat-panel');
  if (panel.style.display === 'none') {
    panel.style.display = 'block';
    if (!document.getElementById('ai-chat-messages').innerHTML) {
      appendAIMsg('ai', '안녕하세요! 루트 최적화 결과에 대해 질문해주세요.\n\n예시:\n- 가장 비싼 루트는?\n- 동부 루트를 합칠 수 있나?\n- 비용 절감 방법은?\n- 이 결과를 요약해줘');
    }
    document.getElementById('ai-chat-input').focus();
  } else {
    panel.style.display = 'none';
  }
}

function appendAIMsg(role, text) {
  var el = document.getElementById('ai-chat-messages');
  var div = document.createElement('div');
  div.style.marginBottom = '10px';
  if (role === 'user') {
    div.innerHTML = '<div style="background:#e8f0fe;padding:8px 12px;border-radius:10px;margin-left:40px;"><b>You:</b> ' + text.replace(/\n/g,'<br>') + '</div>';
  } else {
    div.innerHTML = '<div style="background:#f8f9fa;padding:8px 12px;border-radius:10px;margin-right:20px;"><b><i class="bi bi-robot"></i> AI:</b> ' + text.replace(/\n/g,'<br>') + '</div>';
  }
  el.appendChild(div);
  el.scrollTop = el.scrollHeight;
}

function buildRouteContext() {
  var result = window._lastResult;
  if (!result) return 'No result available.';
  var lines = [];
  lines.push('Route Optimization Results:');
  [result.ms_result, result.l1_result].forEach(function(gr) {
    if (!gr || !gr.routes || gr.routes.length === 0) return;
    lines.push('\\nGroup: ' + gr.group + ' | Origin: ' + (gr.origin||'N/A') + ' | Total Cost: $' + gr.total_cost.toLocaleString() + ' | Routes: ' + gr.routes.length);
    gr.routes.forEach(function(rt, ri) {
      lines.push('  Route ' + (ri+1) + ': POs=' + rt.po_numbers.join(',') + ' | Distance=' + rt.distance_km + 'km | Cost=$' + rt.cost.toLocaleString() + ' | Units=' + rt.total_units);
      rt.schedule.forEach(function(st) {
        lines.push('    Stop' + st.stop + ': DC=' + (st.dc_code||'') + ' ' + st.warehouse + ' | PO=' + st.po_number + ' | Qty=' + st.quantity + ' | ' + st.segment_km + 'km | Arrival=' + st.arrival_time + ' | AdjArrival=' + st.adjusted_arrival);
      });
    });
  });
  var s = result.summary;
  lines.push('\\nSummary: Total POs=' + s.total_pos + ' | Total Routes=' + s.total_routes + ' | Grand Total=$' + s.grand_total.toLocaleString());
  lines.push('Assumptions: Base=$' + result.assumptions.base_cost + '/truck, $' + result.assumptions.cost_per_km + '/km, MaxStops=' + result.assumptions.max_stops + ', MaxCapacity=' + result.assumptions.max_capacity + ' units');
  return lines.join('\\n');
}

async function sendAIChat() {
  var input = document.getElementById('ai-chat-input');
  var q = input.value.trim();
  if (!q) return;
  input.value = '';
  appendAIMsg('user', q);

  // Show typing indicator
  var typing = document.createElement('div');
  typing.id = 'ai-typing';
  typing.innerHTML = '<div style="background:#f8f9fa;padding:8px 12px;border-radius:10px;margin-right:20px;"><i class="bi bi-hourglass-split"></i> 분석 중...</div>';
  document.getElementById('ai-chat-messages').appendChild(typing);

  var context = buildRouteContext();
  try {
    var resp = await fetch('/api/ai-chat', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({question: q, context: context}),
    });
    var data = await resp.json();
    var typingEl = document.getElementById('ai-typing');
    if (typingEl) typingEl.remove();
    if (data.answer) {
      appendAIMsg('ai', data.answer);
    } else {
      appendAIMsg('ai', 'Error: ' + (data.error || 'Unknown error'));
    }
  } catch(e) {
    var typingEl = document.getElementById('ai-typing');
    if (typingEl) typingEl.remove();
    appendAIMsg('ai', 'Error: ' + e.message);
  }
}

function highlightRoute(gi, ri) {
  if (!window._routeLayerGroup) return;

  // Dim all markers except selected route + origin
  window._mapMarkers.forEach(function(m) {
    if (!m.marker) return;
    if ((m.gi === gi && m.ri === ri) || (m.gi === gi && m.ri === -99)) {
      m.marker.setOpacity(1);
    } else {
      m.marker.setOpacity(0.1);
    }
  });

  // Dim all polylines except ones tagged with this route
  window._routeLayerGroup.getLayers().forEach(function(layer) {
    if (layer instanceof L.Polyline && !(layer instanceof L.Polygon)) {
      if (layer._routeTag && layer._routeTag.gi === gi && layer._routeTag.ri === ri) {
        layer.setStyle({opacity: 1});
      } else if (layer._routeTag) {
        layer.setStyle({opacity: 0.08});
      }
    }
  });
}

function showAllRoutes() {
  if (!window._mapMarkers) return;
  window._mapMarkers.forEach(function(m) {
    if (m.marker) m.marker.setOpacity(1);
  });
  if (window._routeLayerGroup) {
    window._routeLayerGroup.getLayers().forEach(function(layer) {
      if (layer instanceof L.Polyline && !(layer instanceof L.Polygon)) {
        layer.setStyle({opacity: 0.8});
      }
    });
  }
  window._hiddenLayers = [];
}

function showRouteModal(gi, ri, color) {
  var result = window._lastResult;
  if (!result) return;
  var gr = [result.ms_result, result.l1_result][gi];
  if (!gr || !gr.routes[ri]) return;
  var rt = gr.routes[ri];
  var a = result.assumptions;
  var distCost = (rt.cost - a.base_cost).toFixed(2);

  var h = '<div style="border-left:4px solid ' + color + ';padding-left:12px;">';
  h += '<div class="d-flex justify-content-between align-items-center mb-2">';
  h += '<h5 class="mb-0" style="color:' + color + ';">' + rt.schedule.map(function(s){return s.dc_code;}).join(' → ') + ' <span class="badge bg-info">' + rt.total_units + '</span> <small class="text-muted">' + rt.po_numbers.join(', ') + '</small></h5>';
  h += '<span class="fs-5 text-success fw-bold">$' + rt.cost.toLocaleString() + '</span></div>';
  h += '<div class="row small mb-2">';
  h += '<div class="col-3"><strong>Pickup:</strong> ' + addDayName(rt.departure) + '</div>';
  h += '<div class="col-3"><strong>Distance:</strong> ' + rt.distance_km.toLocaleString() + ' km</div>';
  h += '<div class="col-3"><strong>Units:</strong> ' + rt.total_units + '</div>';
  h += '<div class="col-3"><strong>Base:</strong> $' + a.base_cost + ' + <strong>Dist:</strong> $' + distCost + '</div></div>';

  h += '<table class="table table-sm table-bordered table-striped small mb-0">';
  h += '<thead class="table-light"><tr><th>Stop</th><th>DC</th><th>Warehouse</th><th>PO</th><th>Qty</th><th>Seg km</th><th>Hours</th><th>마지노 출발</th><th>Arrival</th><th>Adj Arrival</th><th>대기</th><th>Due Date</th><th>Operating Hours</th></tr></thead><tbody>';
  rt.schedule.forEach(function(st) {
    var opHrs = (result.receiving_schedules && st.receiving_code) ? (result.receiving_schedules[st.receiving_code] || '') : '';
    var segHrs = st.segment_hours || (st.segment_km / 80).toFixed(1);
    h += '<tr><td>' + st.stop + '</td>';
    h += '<td><span class="badge bg-secondary">' + (st.dc_code||'') + '</span></td>';
    h += '<td>' + st.warehouse + '</td><td>' + st.po_number + '</td><td>' + st.quantity + '</td>';
    h += '<td>' + st.segment_km + '</td><td>' + segHrs + 'h</td>';
    h += '<td class="text-danger fw-bold">' + addDayName(st.latest_departure||'') + '</td>';
    h += '<td>' + addDayName(st.arrival_time) + '</td>';
    h += '<td>' + addDayName(st.adjusted_arrival) + '</td>';
    var wh2='-';try{var aa=st.arrival_time.split(' '),bb=st.adjusted_arrival.split(' ');var dd1=new Date(aa[0]+'T'+aa[1]),dd2=new Date(bb[0]+'T'+bb[1]);var whh=Math.round((dd2-dd1)/3600000*10)/10;wh2=whh>0?whh+'h':'-';}catch(e){}
    h += '<td>' + wh2 + '</td>';
    h += '<td>' + st.due_date + '</td><td>' + opHrs + '</td></tr>';
  });
  h += '</tbody></table></div>';

  document.getElementById('routeModalBody').innerHTML = h;
  // Hide other routes, show only selected
  highlightRoute(gi, ri);

  var modalEl = document.getElementById('routeDetailModal');
  var modal = new bootstrap.Modal(modalEl);
  modal.show();

  // Restore all routes when modal closes
  modalEl.addEventListener('hidden.bs.modal', function() {
    showAllRoutes();
  }, {once: true});
}

function printMap() {
  var result = window._lastResult;
  if (!result) { alert('결과가 없습니다.'); return; }
  var map = window._mapInstance;
  if (!map) { alert('지도가 없습니다.'); return; }

  // Use static map image from OSM (fast, no capture needed)
  var center = map.getCenter();
  var zoom = map.getZoom();
  var size = map.getSize();
  // Build static map URL with route markers
  var coords = result.coords || {};
  var markerParams = '';
  var routeColors = ['red','blue','green','orange','purple','cyan','chocolate','darkslategray'];
  var colorIdx = 0;
  [result.ms_result, result.l1_result].forEach(function(gr) {
    if (!gr || !gr.routes) return;
    gr.routes.forEach(function(rt) {
      var clr = routeColors[colorIdx % routeColors.length];
      colorIdx++;
      rt.schedule.forEach(function(st) {
        var wc = coords[st.warehouse];
        if (wc) markerParams += '&markers=color:' + clr + '|label:' + (st.dc_code||'') + '|' + wc.lat + ',' + wc.lon;
      });
    });
  });

  // Build compact summary table - all routes in one table per group
  var summaryHtml = '';
  [result.ms_result, result.l1_result].forEach(function(gr) {
    if (!gr || !gr.routes || gr.routes.length === 0) return;
    summaryHtml += '<h3>' + gr.group + ' (' + (gr.origin||'') + ') - $' + Math.round(gr.total_cost).toLocaleString() + ' | ' + gr.routes.length + ' routes</h3>';
    summaryHtml += '<table border="1" cellpadding="2" cellspacing="0" style="border-collapse:collapse;font-size:9px;width:100%;">';
    summaryHtml += '<tr style="background:#003d7a;color:#fff;"><th>Route</th><th>DC</th><th>PO</th><th>Qty</th><th>km</th><th>Cost</th><th>Arrival</th></tr>';
    gr.routes.forEach(function(rt, ri) {
      rt.schedule.forEach(function(st, si) {
        summaryHtml += '<tr' + (si===0?' style="border-top:2px solid #000;"':'') + '>';
        summaryHtml += '<td>' + (si===0?'R'+(ri+1):'') + '</td>';
        summaryHtml += '<td><b>' + (st.dc_code||'') + '</b></td>';
        summaryHtml += '<td>' + st.po_number + '</td>';
        summaryHtml += '<td>' + st.quantity + '</td>';
        summaryHtml += '<td>' + st.segment_km + '</td>';
        summaryHtml += '<td>' + (si===0?'$'+rt.cost.toLocaleString():'') + '</td>';
        summaryHtml += '<td>' + st.arrival_time + '</td></tr>';
      });
    });
    summaryHtml += '</table>';
  });

  // Open print window with embedded map iframe + tables
  var win = window.open('', '_blank');
  var mapHtml = '<div id="print-map" style="width:100%;height:350px;border:1px solid #ddd;border-radius:8px;margin-bottom:10px;"></div>';
  win.document.write('<html><head><title>Route Optimization Report</title>');
  win.document.write('<link href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" rel="stylesheet">');
  win.document.write('<style>body{font-family:Arial,sans-serif;margin:10px;font-size:10px;}h2{font-size:16px;margin:5px 0;}h3{font-size:12px;margin:6px 0 3px;}table{page-break-inside:avoid;margin-bottom:4px;width:100%;}td,th{padding:1px 3px !important;}@page{size:A4 landscape;margin:8mm;}@media print{.no-print{display:none;}body{margin:5mm;}#map-img{width:100%;}}</style></head><body>');
  win.document.write('<h2>Innofoods Route Optimization Report</h2>');
  win.document.write('<div id="print-map" style="width:100%;height:550px;margin-bottom:10px;"></div>');
  win.document.write('<img id="map-img" style="display:none;width:100%;border:1px solid #ddd;border-radius:4px;margin-bottom:10px;">');
  win.document.write(summaryHtml);
  win.document.write('<br><button class="no-print" id="print-btn" disabled style="padding:8px 24px;font-size:14px;cursor:pointer;">지도 로딩 중...</button>');
  win.document.write('<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"><\/script>');
  win.document.write('<script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"><\/script>');
  win.document.write('<script>var routeData=' + JSON.stringify({ms:result.ms_result,l1:result.l1_result,coords:coords}) + ';');
  win.document.write('setTimeout(function(){');
  win.document.write('var m=L.map("print-map",{preferCanvas:true}).setView([' + center.lat + ',' + center.lng + '],' + zoom + ');');
  win.document.write('L.tileLayer("https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png",{attribution:"OSM",maxZoom:18}).addTo(m);');
  win.document.write('var colors=["#e74c3c","#3498db","#2ecc71","#f39c12","#9b59b6","#1abc9c","#e67e22","#34495e"];var ci=0;');
  win.document.write('var d=routeData;[d.ms,d.l1].forEach(function(gr){if(!gr||!gr.routes)return;var isMS=gr.group==="MS-WH";');
  win.document.write('var ok=isMS?"__ms_origin__":"__l1_origin__";var oc=d.coords[ok];');
  win.document.write('if(oc){var oIco=L.divIcon({html:\'<div style="background:\'+(isMS?"#e74c3c":"#3498db")+\';color:#fff;border-radius:50%;width:30px;height:30px;display:flex;align-items:center;justify-content:center;font-weight:bold;font-size:14px;border:2px solid #fff;box-shadow:0 2px 5px rgba(0,0,0,0.3);">&#9733;</div>\',className:"",iconSize:[30,30],iconAnchor:[15,15]});L.marker([oc.lat,oc.lon],{icon:oIco}).addTo(m);}');
  win.document.write('gr.routes.forEach(function(rt){var c=colors[ci%colors.length];ci++;');
  win.document.write('var ll=oc?[[oc.lat,oc.lon]]:[];');
  win.document.write('rt.schedule.forEach(function(st){var wc=d.coords[st.warehouse];if(!wc)return;ll.push([wc.lat,wc.lon]);');
  win.document.write('L.marker([wc.lat,wc.lon],{icon:L.divIcon({html:\'<div style="background:\'+c+\';color:#fff;border-radius:8px;min-width:28px;height:22px;padding:0 4px;display:flex;align-items:center;justify-content:center;font-weight:bold;font-size:10px;border:2px solid #fff;box-shadow:0 2px 4px rgba(0,0,0,0.3);white-space:nowrap;">\'+(st.dc_code||"")+\'</div>\',className:"",iconSize:[28,22],iconAnchor:[14,11]})}).addTo(m);});');
  win.document.write('if(ll.length>1)L.polyline(ll,{color:c,weight:3}).addTo(m);});});');
  // After tiles load, convert map to image for clean printing
  win.document.write('setTimeout(function(){');
  win.document.write('var mapEl=document.getElementById("print-map");');
  win.document.write('var canvases=mapEl.querySelectorAll("canvas");');
  win.document.write('html2canvas(mapEl,{useCORS:true,allowTaint:true,scale:2,onclone:function(doc){var oc=mapEl.querySelectorAll("canvas");var cc=doc.getElementById("print-map").querySelectorAll("canvas");for(var i=0;i<oc.length;i++){if(cc[i]){cc[i].getContext("2d").drawImage(oc[i],0,0);}}}}).then(function(canvas){');
  win.document.write('var img=document.getElementById("map-img");img.src=canvas.toDataURL("image/png");img.style.display="block";');
  win.document.write('mapEl.style.display="none";');
  win.document.write('var btn=document.getElementById("print-btn");btn.disabled=false;btn.textContent="인쇄";btn.onclick=function(){window.print();};');
  win.document.write('});},2000);');
  win.document.write('},500);<\/script></body></html>');
  win.document.close();
}

function backToInput() {
  window._mapInstance = null; window._routeLayerGroup = null;
  document.getElementById('results-section').style.display = 'none';
  document.getElementById('input-section').style.display = 'block';
}

function exportJSON() {
  if (!window._lastResult) return;
  const blob = new Blob([JSON.stringify(window._lastResult, null, 2)], {type: 'application/json'});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = 'route_optimization_result.json';
  a.click();
}

async function uploadExcel(input) {
  if (!input.files || !input.files[0]) return;
  var formData = new FormData();
  formData.append('file', input.files[0]);
  try {
    var resp = await fetch('/api/upload-excel', { method: 'POST', body: formData });
    var data = await resp.json();
    if (data.success && data.pos && data.pos.length > 0) {
      clearAll();
      // Reverse-map warehouse -> DC for display
      var whToDc = {};
      for (var dc in DC_MAP) { whToDc[DC_MAP[dc]] = dc; }
      for (var i = 0; i < data.pos.length; i++) {
        var p = data.pos[i];
        if (!p.dc && p.warehouse) p.dc = whToDc[p.warehouse] || '';
        addRow(p);
      }
      var msg = data.pos.length + '건 로드 완료';
      if (data.warnings && data.warnings.length > 0) {
        msg += '\n\n⚠ Validation 경고:\n' + data.warnings.join('\n');
      }
      alert(msg);
    } else {
      alert('데이터가 없거나 파싱 실패: ' + (data.error || ''));
    }
  } catch(e) {
    alert('업로드 실패: ' + e.message);
  }
  input.value = '';
}

// Keyboard shortcuts: 1=cut mode hint, 2=merge, 3=clear
document.addEventListener('keydown', function(e) {
  // F4 and ESC work anytime
  if (e.key === 'F4') { e.preventDefault(); toggleFullscreen(); return; }
  if (e.key === 'Escape' && window._isFullscreen) { exitFullscreen(); return; }
  // A, D only in fullscreen
  if (!window._isFullscreen) return;
  if (e.target.tagName === 'INPUT' || e.target.tagName === 'TEXTAREA' || e.target.tagName === 'SELECT') return;
  if (e.key === 'a' || e.key === 'A') { e.preventDefault(); mapMerge(); }
  else if (e.key === 'd' || e.key === 'D') { e.preventDefault(); cutSelectedSegment(); }
});


function toggleFullscreen() {
  if (window._isFullscreen) {
    exitFullscreen();
  } else {
    var container = document.getElementById('route-map').parentElement;
    container.style.cssText = 'position:fixed;top:0;left:0;width:100vw;height:100vh;z-index:10000;margin:0;border-radius:0;padding:0;';
    var mapEl = document.getElementById('route-map');
    mapEl.style.height = '100vh';
    mapEl.style.borderRadius = '0';
    window._isFullscreen = true;
    document.getElementById('fs-icon').className = 'bi bi-fullscreen-exit';
    setTimeout(function() {
      if (window._mapInstance) window._mapInstance.invalidateSize();
    }, 100);
  }
}

function exitFullscreen() {
  var container = document.getElementById('route-map').parentElement;
  // Reset all styles on the relative container
  container.style.cssText = 'position:relative;';
  var mapEl = document.getElementById('route-map');
  mapEl.style.height = '75vh';
  mapEl.style.minHeight = '500px';
  mapEl.style.borderRadius = '8px';
  window._isFullscreen = false;
  document.getElementById('fs-icon').className = 'bi bi-arrows-fullscreen';
  // Delay invalidateSize to let DOM settle
  setTimeout(function() {
    if (window._mapInstance) window._mapInstance.invalidateSize();
  }, 100);
}

function showError(msg) {
  console.error('showError:', msg);
  if (window._isFullscreen) {
    // In fullscreen, modal might be hidden. Use a toast-like overlay instead.
    var toast = document.createElement('div');
    toast.style.cssText = 'position:fixed;top:50%;left:50%;transform:translate(-50%,-50%);z-index:20000;background:#dc3545;color:#fff;padding:16px 24px;border-radius:12px;font-size:14px;font-weight:bold;box-shadow:0 4px 20px rgba(0,0,0,.3);max-width:400px;text-align:center;';
    toast.textContent = msg;
    document.body.appendChild(toast);
    setTimeout(function() { toast.remove(); }, 4000);
  } else {
    var el = document.getElementById('errorModalBody');
    if (el) {
      el.textContent = msg;
      new bootstrap.Modal(document.getElementById('errorModal')).show();
    } else {
      alert(msg);
    }
  }
}
function cutSelectedSegment() {
  if (!window._selectedSeg) { showError('끊을 선을 먼저 클릭하세요.'); return; }
  var s = window._selectedSeg;
  window._selectedSeg = null;
  mapCutSegment(s.gi, s.ri, s.seg);
}

// Init
addRow();
</script>
</body>
</html>
"""


# ---------------------------------------------------------------------------
# Flask Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    origins = sorted(ORIGIN_COORDS.keys())
    ms_whs = [w for w in WAREHOUSES if w["group"] == "MS-WH"]
    l1_whs = [w for w in WAREHOUSES if w["group"] == "L1-WH"]
    # Build reverse map: warehouse name -> DC code
    wh_to_dc = {v: k for k, v in DC_CODE_MAP.items()}
    # Add dc_code to sample data for display
    sample_with_dc = []
    for po in SAMPLE_POS:
        d = dict(po)
        d["dc"] = wh_to_dc.get(po["warehouse"], "")
        sample_with_dc.append(d)

    return render_template_string(
        MAIN_PAGE,
        origins=origins,
        default_ms=DEFAULT_MS_ORIGIN,
        default_l1=DEFAULT_L1_ORIGIN,
        warehouses_json=json.dumps(WAREHOUSES),
        sample_json=json.dumps(sample_with_dc),
        dc_map_json=json.dumps(DC_CODE_MAP),
        ms_warehouses=ms_whs,
        l1_warehouses=l1_whs,
    )


@app.route("/api/warehouses", methods=["GET"])
def api_warehouses():
    """Return warehouse list as JSON."""
    return jsonify(WAREHOUSES)


@app.route("/api/optimize", methods=["POST"])
def api_optimize():
    """
    Accept JSON POST with:
      { "pos": [...], "ms_origin": "...", "l1_origin": "..." }
    Return optimization results as JSON.
    """
    data = request.get_json(force=True)
    if not data:
        return jsonify({"success": False, "errors": ["No JSON body provided"]}), 400

    pos_data = data.get("pos", [])
    if not pos_data:
        return jsonify({"success": False, "errors": ["No PO data provided"]}), 400

    ms_origin_str = data.get("ms_origin", DEFAULT_MS_ORIGIN)
    l1_origin_str = data.get("l1_origin", DEFAULT_L1_ORIGIN)
    balance_weight = float(data.get("balance_weight", 0.0))

    result = run_optimization(pos_data, ms_origin_str, l1_origin_str, balance_weight)
    status = 200 if result.get("success") else 400
    return jsonify(result), status


@app.route("/api/recalc-route", methods=["POST"])
def api_recalc_route():
    """Recalculate a single route with a new pickup time."""
    from route_optimizer import evaluate_route, load_cache, COORDS, ORIGIN_COORDS, ORIGIN_TZ, AVG_SPEED_KMH
    data = request.get_json(force=True)
    po_numbers = data.get("po_numbers", [])
    warehouses = data.get("warehouses", [])
    quantities = data.get("quantities", [])
    due_dates = data.get("due_dates", [])
    pickup_str = data.get("pickup_time", "")
    group = data.get("group", "MS-WH")

    try:
        pickup_dt = datetime.strptime(pickup_str, "%Y-%m-%d %H:%M")
    except ValueError:
        return jsonify({"success": False, "error": "Invalid pickup_time format"})

    # Determine origin
    if group == "MS-WH":
        origin_name = DEFAULT_MS_ORIGIN
    else:
        origin_name = DEFAULT_L1_ORIGIN
    origin_coord = ORIGIN_COORDS.get(origin_name)
    if not origin_coord:
        return jsonify({"success": False, "error": f"Unknown origin: {origin_name}"})

    # Build PO list
    po_list = []
    for i in range(len(po_numbers)):
        po_list.append({
            "po_number": po_numbers[i],
            "warehouse": warehouses[i] if i < len(warehouses) else "",
            "quantity": quantities[i] if i < len(quantities) else 0,
            "due_date": due_dates[i] if i < len(due_dates) else "2099-12-31",
            "inventory_available_date": pickup_str[:10],
        })

    import itertools
    from zoneinfo import ZoneInfo
    from route_optimizer import save_cache
    o_tz = ORIGIN_TZ.get(origin_name, ZoneInfo("America/New_York"))
    cache = load_cache()

    # Try all permutations to find best order
    best_route = None
    if len(po_list) <= 6:
        for perm in itertools.permutations(range(len(po_list))):
            ordered = [po_list[j] for j in perm]
            route = evaluate_route(origin_coord, ordered, pickup_dt, cache, o_tz)
            if best_route is None or route["cost"] < best_route["cost"]:
                best_route = route
    else:
        best_route = evaluate_route(origin_coord, po_list, pickup_dt, cache, o_tz)

    save_cache(cache)
    return jsonify({"success": True, "route": best_route})


@app.route("/api/auto-schedule", methods=["POST"])
def api_auto_schedule():
    """Auto-determine pickup dates for all routes."""
    from route_optimizer import (auto_schedule_route, load_cache, save_cache,
                                  ORIGIN_COORDS, ORIGIN_TZ, COORDS, evaluate_route)
    from zoneinfo import ZoneInfo

    data = request.get_json(force=True)
    routes = data.get("routes", [])
    group = data.get("group", "MS-WH")

    if group == "MS-WH":
        origin_name = DEFAULT_MS_ORIGIN
    else:
        origin_name = DEFAULT_L1_ORIGIN

    origin_coord = ORIGIN_COORDS.get(origin_name)
    o_tz = ORIGIN_TZ.get(origin_name, ZoneInfo("America/New_York"))
    cache = load_cache()

    from route_optimizer import distribute_pickup_times, balance_departure_dates

    results = []
    for rt in routes:
        sched = auto_schedule_route(rt, origin_coord, o_tz, cache)
        results.append(sched)

    # Distribute pickup times so same-day routes get different hours
    results = balance_departure_dates(results)
    results = distribute_pickup_times(results)

    # Re-simulate routes with actual departure times (not pickup)
    for i, sched in enumerate(results):
        if not sched.get("departure"):
            continue
        depart_dt = datetime.strptime(sched["departure"], "%Y-%m-%d %H:%M")
        route_stops = routes[i].get("schedule", [])
        sim = evaluate_route(origin_coord,
            [{"warehouse": st["warehouse"], "po_number": st["po_number"],
              "quantity": st["quantity"], "due_date": st["due_date"],
              "inventory_available_date": (depart_dt - timedelta(days=1)).strftime("%Y-%m-%d")}
             for st in route_stops],
            depart_dt, cache, o_tz)
        results[i]["simulated_route"] = sim

    save_cache(cache)
    return jsonify({"success": True, "schedules": results})


@app.route("/api/learn", methods=["POST"])
def api_learn():
    """Save current route result as a learned preference."""
    from route_optimizer import learn_from_result, load_preferences
    data = request.get_json(force=True)
    routes = data.get("routes", [])
    group = data.get("group", "")

    if not routes:
        return jsonify({"success": False, "error": "No routes provided"})

    learn_from_result(routes, group)
    prefs = load_preferences()
    pair_count = len(prefs.get("pair_scores", {}))
    saved_count = len(prefs.get("saved_results", []))

    return jsonify({
        "success": True,
        "message": f"학습 완료! 저장된 패턴: {pair_count}개, 히스토리: {saved_count}건",
        "pair_scores": prefs.get("pair_scores", {}),
    })


@app.route("/api/preferences", methods=["GET"])
def api_preferences():
    """Get current learned preferences."""
    from route_optimizer import load_preferences
    return jsonify(load_preferences())


@app.route("/api/ai-chat", methods=["POST"])
def api_ai_chat():
    """Send route data + question to Ollama for analysis."""
    import requests as req
    data = request.get_json(force=True)
    question = data.get("question", "")
    context = data.get("context", "")

    system_prompt = """You are a logistics route optimization analyst. You analyze delivery route data and provide insights in Korean.
You help with:
- Route cost analysis and comparison
- Suggestions for route improvements
- Identifying inefficient routes
- Answering questions about delivery schedules and distances
Keep answers concise and practical. Use numbers and data from the context."""

    user_prompt = f"""다음은 현재 루트 최적화 결과입니다:

{context}

질문: {question}"""

    try:
        resp = req.post("http://localhost:11434/api/generate", json={
            "model": "qwen2.5:14b-instruct",
            "prompt": user_prompt,
            "system": system_prompt,
            "stream": False,
            "options": {"temperature": 0.3, "num_predict": 1024},
        }, timeout=60)
        result = resp.json()
        answer = result.get("response", "").strip()
        if not answer:
            return jsonify({"error": "Empty response from AI"})
        return jsonify({"answer": answer})
    except req.exceptions.Timeout:
        return jsonify({"error": "AI 응답 시간 초과 (60초)"})
    except Exception as e:
        return jsonify({"error": str(e)})


@app.route("/api/sample", methods=["GET"])
def api_sample():
    """Return sample PO data."""
    return jsonify(SAMPLE_POS)


@app.route("/api/template.xlsx", methods=["GET"])
def download_template():
    """Download PO sample Excel template."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "PO Data"

    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="003D7A", end_color="003D7A", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    headers = [("SHIP TO", 45), ("DC", 8), ("PO #", 22), ("ITEM", 10),
               ("QTY (PALLETS)", 14), ("Revised QTY", 12), ("QTY (UNITS)", 13),
               ("REQ DELIVERY", 15), ("ACTUAL DELIVERY", 16)]

    col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    for col, (name, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border
        ws.column_dimensions[col_letters[col-1]].width = width

    # Sample data matching real PO format
    samples = [
        ("MCLANE SUNWEST\n14149 WEST MCDOWELL ROAD\nGOODYEAR, AZ 85395-25", "MS", "MS10073402-01", "DSKT", 58, 24, 15776, "3/24", ""),
        ("MCLANE WESTERN\n2100 E KEN PRATT BLVD\nLONGMONT, CO 80504", "MW", "MW10094380-01", "DSKT", 100, 44, 27200, "3/24", ""),
        ("MCLANE HIGH PLAINS\n1717 E LOOP 289\nLUBBOCK, TX 79403", "HP", "HP10067649-01", "DSKT", 33, 14, 8976, "3/24", ""),
        ("MCLANE NORTH TEXAS\n7550 OAK GROVE RD\nFORT WORTH, TX 76140", "NT", "NT10187696-01", "DSKT", 88, 38, 23936, "3/24", ""),
        ("MCLANE NORTHWEST\n9611 45TH AVE SW BLDG 4\nLAKEWOOD, WA 98499", "NW", "NW10089277-01", "DSKT", 10, 4, 2720, "3/24", ""),
        ("MCLANE SO. CALIF.\n4472 GEORGIA BLVD\nSan Bernardino, CA 92407-18", "SZ", "SZ10081334-01", "DSKT", 47, 20, 12784, "3/24", ""),
        ("MCLANE SOUTHWEST\n2828 INDUSTRIAL BLVD\nTEMPLE, TX 76504", "SW", "SW10095010-01", "DSKT", 77, 34, 20944, "3/24", ""),
        ("MCLANE MINNESOTA\n1111 WEST 5TH STREET\nNORTHFIELD, MN 55057", "MN", "MN10098698-01", "DSKT", 46, 20, 12512, "3/24", ""),
        ("MCLANE PACIFIC\n3876 CHILDS AVE\nMERCED, CA 95340", "MP", "MP10072671-01", "DSKT", 31, 12, 8432, "3/24", ""),
        ("MCLANE CAROLINA\n7253 NC-4\nBATTLEBORO, NC 27809", "NC", "NC10097501-01", "DSKT", 68, 30, 18496, "3/24", ""),
        ("MCLANE MID-ATLANTIC\n56 MCLANE DRIVE\nFredricksburg, VA 22406", "MZ", "MZ10084952-01", "DSKT", 28, 12, 7616, "3/24", ""),
        ("MCLANE NORTHEAST\n2828 MCLANE DRIVE\nBALDWINSVILLE, NY 13027", "NE", "NE10085348-01", "DSKT", 35, 14, 9520, "3/24", ""),
        ("MCLANE PA\n43 VALLY VIEW BUSINESS PA\nJESSUP, PA 18434", "PA", "PA10086419-01", "DSKT", 46, 20, 12512, "3/24", ""),
        ("MCLANE NE/CONCORD\n932 MAPLE ST\nCONTOOCOOK, NH 03229", "MY", "MY10075692-01", "DSKT", 15, 6, 4080, "3/24", ""),
        ("MCLANE DOTHAN\n100 MCLANE PARKWAY\nCOTTONWOOD, AL 36320", "MD", "MD10087384-01", "DSKT", 20, 10, 5440, "3/24", ""),
        ("MCLANE SOUTHERN\n2104 MANUFACTURERS BLVD\nBROOKHAVEN, MS 39601", "SO", "SO10081532-01", "DSKT", 58, 24, 15776, "3/24", ""),
        ("MCLANE SOUTHEAST\n300 NORTH HWY 29\nATHENS, GA 30601", "SE", "SE10198234-01", "DSKT", 80, 36, 21760, "3/24", ""),
        ("MCLANE SUNEAST\n1818 POINCIANA BLVD\nKISSIMMEE, FL 34758", "ME", "ME10132021-01", "DSKT", 30, 12, 8160, "3/24", ""),
        ("MCLANE OCALA\n910 NW 50TH AVE\nOCALA, FL 34482", "FE", "FE10190732-01", "DSKT", 20, 8, 5440, "3/24", ""),
        ("MCLANE OZARK\n2788 E SAWYER ROAD\nREPUBLIC, MO 65738", "MO", "MO10096866-01", "DSKT", 128, 58, 34816, "3/24", ""),
        ("MCLANE CUMBERLAND\n104 MCLANE BLVD\nNICHOLASVILLE, KY 40356", "MK", "MK10090088-01", "DSKT", 48, 20, 13056, "3/24", ""),
        ("MCLANE OHIO\n3200 MCLANE DRIVE\nFINDLAY, OH 45840", "MG", "MG10189076-01", "DSKT", 94, 40, 25568, "3/24", ""),
    ]
    wrap_align = Alignment(vertical="top", wrap_text=True)
    num_fmt = '#,##0'
    for r, row_data in enumerate(samples, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            if c == 2:  # SHIP TO - wrap text
                cell.alignment = wrap_align
        ws.row_dimensions[r].height = 45

    # Warehouse list sheet
    ws2 = wb.create_sheet("Warehouse List")
    ws2_headers = [("DC Code", 10), ("Warehouse Name", 35), ("City", 20), ("State", 18), ("Group", 10)]
    ws2_letters = ['A', 'B', 'C', 'D', 'E']
    for col, (name, width) in enumerate(ws2_headers, 1):
        cell = ws2.cell(row=1, column=col, value=name)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border
        ws2.column_dimensions[ws2_letters[col-1]].width = width

    for r, wh in enumerate(WAREHOUSES, 2):
        ws2.cell(row=r, column=1, value=wh["receiving_code"]).border = border
        ws2.cell(row=r, column=2, value=wh["name"]).border = border
        ws2.cell(row=r, column=3, value=wh["city"]).border = border
        ws2.cell(row=r, column=4, value=wh["state"]).border = border
        ws2.cell(row=r, column=5, value=wh["group"]).border = border

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=po_template.xlsx"},
    )


_STATE_ABBREVS = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "Florida": "FL", "Georgia": "GA", "Hawaii": "HI", "Idaho": "ID",
    "Illinois": "IL", "Indiana": "IN", "Iowa": "IA", "Kansas": "KS",
    "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME", "Maryland": "MD",
    "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS",
    "Missouri": "MO", "Montana": "MT", "Nebraska": "NE", "Nevada": "NV",
    "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM", "New York": "NY",
    "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK",
    "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI", "South Carolina": "SC",
    "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX", "Utah": "UT",
    "Vermont": "VT", "Virginia": "VA", "Washington": "WA", "West Virginia": "WV",
    "Wisconsin": "WI", "Wyoming": "WY",
}

def _state_abbrev(state_name: str) -> str:
    return _STATE_ABBREVS.get(state_name, state_name.upper()[:2])


@app.route("/api/upload-excel", methods=["POST"])
def upload_excel():
    """Parse uploaded Excel file and return PO data as JSON."""
    from openpyxl import load_workbook
    from io import BytesIO

    if "file" not in request.files:
        return jsonify({"success": False, "error": "No file uploaded"}), 400

    f = request.files["file"]
    wb = load_workbook(BytesIO(f.read()), data_only=True)
    ws = wb.active

    # Read header row
    headers = [str(cell.value or "").strip() for cell in ws[1]]

    # Map expected columns - support both standard format and real PO format
    col_map = {}
    is_real_format = False  # True if using DC/PO#/Revised QTY columns
    expected = {
        "po number": "po_number",
        "warehouse": "warehouse",
        "due date": "due_date",
        "quantity": "quantity",
        "product type": "product_type",
        "inventory available date": "inventory_available_date",
    }
    for i, h in enumerate(headers):
        key = h.lower().strip()
        if key in expected:
            col_map[expected[key]] = i
        # Standard format partial matches
        elif "po" in key and ("number" in key or "#" in key):
            col_map["po_number"] = i
        elif key == "po #" or key == "po#":
            col_map["po_number"] = i
        elif "warehouse" in key:
            col_map["warehouse"] = i
        elif "due" in key:
            col_map["due_date"] = i
        elif "product" in key or key == "item":
            col_map["product_type"] = i
        elif "inventory" in key or "available" in key:
            col_map["inventory_available_date"] = i
        # Real PO format columns
        elif key == "dc":
            col_map["dc"] = i
            is_real_format = True
        elif "revised" in key and "qty" in key:
            col_map["revised_qty"] = i  # Track separately
            col_map["quantity"] = i     # Primary
        elif "qty" in key and "unit" in key:
            pass  # skip QTY (UNITS)
        elif "qty" in key and "pallet" in key:
            col_map["qty_pallets"] = i  # Track separately
            if "quantity" not in col_map:
                col_map["quantity"] = i
        elif ("qty" in key or "quantity" in key) and "quantity" not in col_map:
            col_map["quantity"] = i
        elif key == "ship to":
            col_map["ship_to"] = i
        elif "req" in key and "delivery" in key:
            col_map["due_date"] = i
        elif "actual" in key and "delivery" in key:
            pass  # skip ACTUAL DELIVERY
        elif "delivery" in key and "due_date" not in col_map:
            col_map["due_date"] = i

    pos = []
    warnings = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        vals = [cell.value for cell in row]
        if not any(vals):
            continue

        po = {}
        for field, idx in col_map.items():
            val = vals[idx] if idx < len(vals) else None
            if val is None:
                val = ""
            # Handle date objects
            if field in ("due_date", "inventory_available_date"):
                if hasattr(val, "strftime"):
                    val = val.strftime("%Y-%m-%d")
                else:
                    val = str(val).strip()
                    # Handle M/D format like "3/24" -> "2026-03-24"
                    if "/" in val and len(val) <= 5:
                        parts = val.split("/")
                        if len(parts) == 2:
                            try:
                                m, d = int(parts[0]), int(parts[1])
                                val = f"2026-{m:02d}-{d:02d}"
                            except ValueError:
                                pass
                    elif len(val) > 10:
                        val = val[:10]
            elif field in ("quantity", "revised_qty", "qty_pallets"):
                try:
                    val = int(float(str(val).replace(",", "")))
                except (ValueError, TypeError):
                    val = 0
            else:
                val = str(val).strip()
            po[field] = val

        # Fallback: if revised_qty is empty/0, use qty_pallets
        if po.get("quantity", 0) == 0 and "qty_pallets" in po:
            try:
                po["quantity"] = int(float(str(po["qty_pallets"]).replace(",", "")))
            except (ValueError, TypeError):
                pass
        # Clean up temp fields
        po.pop("revised_qty", None)
        po.pop("qty_pallets", None)

        # Resolve DC code -> warehouse name
        dc = ""
        ship_to = po.get("ship_to", "")
        if "dc" in po and po["dc"]:
            dc = po.pop("dc").strip().upper()
            po["warehouse"], po["dc_key"] = _resolve_dc(dc, po.get("po_number", "?"), warnings, ship_to)

        # Fallback: resolve from PO prefix (e.g., MS10073402-01 -> DC: MS)
        if not po.get("warehouse") and po.get("po_number"):
            import re
            m = re.match(r'^([A-Z]{2})', po["po_number"])
            if m:
                dc = m.group(1)
                wh_name, dc_key = _resolve_dc(dc, po.get("po_number", "?"), warnings, ship_to)
                if wh_name:
                    po["warehouse"] = wh_name
                    po["dc_key"] = dc_key

        # Remove ship_to (reference only, not used for validation)
        po.pop("ship_to", None)

        # Set inventory_available_date to 1 week before due_date
        if po.get("due_date"):
            try:
                dd = datetime.strptime(po["due_date"][:10], "%Y-%m-%d")
                po["inventory_available_date"] = (dd - timedelta(days=7)).strftime("%Y-%m-%d")
            except ValueError:
                from datetime import date
                po["inventory_available_date"] = date.today().strftime("%Y-%m-%d")
        elif not po.get("inventory_available_date"):
            from datetime import date
            po["inventory_available_date"] = date.today().strftime("%Y-%m-%d")
        if not po.get("pickup_time"):
            po["pickup_time"] = po["inventory_available_date"] + "T06:00"

        # Set dc to the map key for correct dropdown selection
        if po.get("dc_key"):
            po["dc"] = po.pop("dc_key")

        if po.get("po_number"):
            pos.append(po)

    return jsonify({"success": True, "pos": pos, "warnings": warnings})


# ---------------------------------------------------------------------------
# Warehouse Management Page
# ---------------------------------------------------------------------------
WH_MGMT_PAGE = r"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Warehouse Management - Innofoods Route Optimizer</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css" rel="stylesheet">
<style>
  :root { --cj-blue: #003d7a;
  .modal-backdrop { z-index: 10005 !important; }
  .modal { z-index: 10010 !important; } }
  body { background: #f4f6f9; font-size: 0.88rem; }
  .navbar { background: var(--cj-blue) !important; }
  .card { border: none; box-shadow: 0 1px 4px rgba(0,0,0,.08); }
  .btn-cj { background: var(--cj-blue); color: #fff; border: none; }
  .btn-cj:hover { background: #00274d; color: #fff; }
  .nav-tabs .nav-link.active { font-weight: bold; }
  .table td, .table th { vertical-align: middle; }
</style>
</head>
<body>
<nav class="navbar navbar-dark mb-4">
  <div class="container-fluid">
    <span class="navbar-brand mb-0 h1"><i class="bi bi-building"></i> Warehouse & Schedule Management</span>
    <a href="/" class="btn btn-outline-light btn-sm"><i class="bi bi-arrow-left"></i> Route Optimizer</a>
  </div>
</nav>

<div class="container-fluid" style="max-width:1400px;">

  <ul class="nav nav-tabs mb-3" id="mgmtTabs">
    <li class="nav-item"><a class="nav-link active" data-bs-toggle="tab" href="#tab-wh"><i class="bi bi-building"></i> Warehouses (창고)</a></li>
    <li class="nav-item"><a class="nav-link" data-bs-toggle="tab" href="#tab-sched"><i class="bi bi-clock"></i> Receiving Schedules (운영시간)</a></li>
    <li class="nav-item"><a class="nav-link" data-bs-toggle="tab" href="#tab-origin"><i class="bi bi-geo-alt"></i> Origins (출발지)</a></li>
  </ul>

  <div class="tab-content">

    <!-- Warehouses Tab -->
    <div class="tab-pane fade show active" id="tab-wh">
      <div class="card p-3">
        <div class="d-flex justify-content-between align-items-center mb-2">
          <h6 class="mb-0">Destination Warehouses (목적지 창고)</h6>
          <button class="btn btn-sm btn-cj" onclick="showWhModal()"><i class="bi bi-plus-lg"></i> Add Warehouse</button>
        </div>
        <div class="table-responsive">
          <table class="table table-sm table-bordered table-striped" id="wh-table">
            <thead class="table-dark">
              <tr><th>DC</th><th>Name</th><th>Address</th><th>City</th><th>State</th><th>Zip</th><th>Group</th><th>Lat</th><th>Lon</th><th style="width:90px;">Actions</th></tr>
            </thead>
            <tbody id="wh-tbody"></tbody>
          </table>
        </div>
      </div>
    </div>

    <!-- Schedules Tab -->
    <div class="tab-pane fade" id="tab-sched">
      <div class="card p-3">
        <div class="d-flex justify-content-between align-items-center mb-2">
          <h6 class="mb-0">Receiving Schedules (수령 운영시간)</h6>
          <button class="btn btn-sm btn-cj" onclick="showSchedModal()"><i class="bi bi-plus-lg"></i> Add Schedule</button>
        </div>
        <div class="table-responsive">
          <table class="table table-sm table-bordered table-striped" id="sched-table">
            <thead class="table-dark">
              <tr><th>Code</th><th>Schedule (요일 시간)</th><th>Used By</th><th style="width:90px;">Actions</th></tr>
            </thead>
            <tbody id="sched-tbody"></tbody>
          </table>
        </div>
      </div>
    </div>

    <!-- Origins Tab -->
    <div class="tab-pane fade" id="tab-origin">
      <div class="card p-3">
        <div class="d-flex justify-content-between align-items-center mb-2">
          <h6 class="mb-0">Origin Warehouses (출발지)</h6>
          <button class="btn btn-sm btn-cj" onclick="showOriginModal()"><i class="bi bi-plus-lg"></i> Add Origin</button>
        </div>
        <div class="table-responsive">
          <table class="table table-sm table-bordered table-striped" id="origin-table">
            <thead class="table-dark">
              <tr><th>Name</th><th>Lat</th><th>Lon</th><th style="width:90px;">Actions</th></tr>
            </thead>
            <tbody id="origin-tbody"></tbody>
          </table>
        </div>
      </div>
    </div>

  </div>
</div>

<!-- Warehouse Modal -->
<div class="modal fade" id="whModal" tabindex="-1"><div class="modal-dialog">
  <div class="modal-content">
    <div class="modal-header"><h5 class="modal-title" id="whModalTitle">Add Warehouse</h5><button type="button" class="btn-close" data-bs-dismiss="modal"></button></div>
    <div class="modal-body">
      <input type="hidden" id="wh-edit-idx" value="-1">
      <div class="row g-2">
        <div class="col-4"><label class="form-label small">DC Code</label><input class="form-control form-control-sm" id="wh-dc"></div>
        <div class="col-8"><label class="form-label small">Name</label><input class="form-control form-control-sm" id="wh-name"></div>
        <div class="col-12"><label class="form-label small">Address</label><input class="form-control form-control-sm" id="wh-address"></div>
        <div class="col-5"><label class="form-label small">City</label><input class="form-control form-control-sm" id="wh-city"></div>
        <div class="col-4"><label class="form-label small">State</label><input class="form-control form-control-sm" id="wh-state"></div>
        <div class="col-3"><label class="form-label small">Zip</label><input class="form-control form-control-sm" id="wh-zip"></div>
        <div class="col-4"><label class="form-label small">Group</label>
          <select class="form-select form-select-sm" id="wh-group"><option>MS-WH</option><option>L1-WH</option></select>
        </div>
        <div class="col-8"></div>
        <div class="col-6"><label class="form-label small">Latitude</label><input type="number" step="any" class="form-control form-control-sm" id="wh-lat"></div>
        <div class="col-6"><label class="form-label small">Longitude</label><input type="number" step="any" class="form-control form-control-sm" id="wh-lon"></div>
      </div>
    </div>
    <div class="modal-footer">
      <button class="btn btn-secondary btn-sm" data-bs-dismiss="modal">Cancel</button>
      <button class="btn btn-cj btn-sm" onclick="saveWh()">Save</button>
    </div>
  </div>
</div></div>

<!-- Schedule Modal -->
<div class="modal fade" id="schedModal" tabindex="-1"><div class="modal-dialog">
  <div class="modal-content">
    <div class="modal-header"><h5 class="modal-title" id="schedModalTitle">Add Schedule</h5><button type="button" class="btn-close" data-bs-dismiss="modal"></button></div>
    <div class="modal-body">
      <input type="hidden" id="sched-edit-code" value="">
      <div class="mb-2"><label class="form-label small">Code</label><input class="form-control form-control-sm" id="sched-code"></div>
      <div id="sched-windows">
        <label class="form-label small">Windows (요일 시작-종료)</label>
      </div>
      <button class="btn btn-sm btn-outline-secondary mt-1" onclick="addSchedWindow()"><i class="bi bi-plus"></i> Add Window</button>
    </div>
    <div class="modal-footer">
      <button class="btn btn-secondary btn-sm" data-bs-dismiss="modal">Cancel</button>
      <button class="btn btn-cj btn-sm" onclick="saveSched()">Save</button>
    </div>
  </div>
</div></div>

<!-- Origin Modal -->
<div class="modal fade" id="originModal" tabindex="-1"><div class="modal-dialog">
  <div class="modal-content">
    <div class="modal-header"><h5 class="modal-title" id="originModalTitle">Add Origin</h5><button type="button" class="btn-close" data-bs-dismiss="modal"></button></div>
    <div class="modal-body">
      <input type="hidden" id="origin-edit-name" value="">
      <div class="mb-2"><label class="form-label small">Name</label><input class="form-control form-control-sm" id="origin-name"></div>
      <div class="row g-2">
        <div class="col-6"><label class="form-label small">Latitude</label><input type="number" step="any" class="form-control form-control-sm" id="origin-lat"></div>
        <div class="col-6"><label class="form-label small">Longitude</label><input type="number" step="any" class="form-control form-control-sm" id="origin-lon"></div>
      </div>
    </div>
    <div class="modal-footer">
      <button class="btn btn-secondary btn-sm" data-bs-dismiss="modal">Cancel</button>
      <button class="btn btn-cj btn-sm" onclick="saveOrigin()">Save</button>
    </div>
  </div>
</div></div>

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>
<script>
var DATA = {};
var DC_MAP = {};

function loadData() {
  fetch('/api/warehouse-data').then(r => r.json()).then(d => {
    DATA = d;
    DC_MAP = d.dc_map || {};
    renderAll();
  });
}

function renderAll() {
  renderWarehouses();
  renderSchedules();
  renderOrigins();
}

// ---- Warehouses ----
function renderWarehouses() {
  var tbody = document.getElementById('wh-tbody');
  var html = '';
  for (var i = 0; i < DATA.warehouses.length; i++) {
    var w = DATA.warehouses[i];
    var dc = w.dc_code || '';
    html += '<tr>' +
      '<td>' + dc + '</td>' +
      '<td>' + w.name + '</td>' +
      '<td>' + (w.address || '') + '</td>' +
      '<td>' + w.city + '</td>' +
      '<td>' + w.state + '</td>' +
      '<td>' + w.zip + '</td>' +
      '<td><span class="badge ' + (w.group==='MS-WH'?'bg-danger':'bg-primary') + '">' + w.group + '</span></td>' +
      '<td>' + (w.lat||'') + '</td>' +
      '<td>' + (w.lon||'') + '</td>' +
      '<td>' +
        '<button class="btn btn-sm btn-outline-primary py-0 px-1 me-1" onclick="editWh(' + i + ')"><i class="bi bi-pencil"></i></button>' +
        '<button class="btn btn-sm btn-outline-danger py-0 px-1" onclick="deleteWh(' + i + ')"><i class="bi bi-trash"></i></button>' +
      '</td></tr>';
  }
  tbody.innerHTML = html;
}

function showWhModal(idx) {
  var isEdit = (idx !== undefined && idx >= 0);
  document.getElementById('whModalTitle').textContent = isEdit ? 'Edit Warehouse' : 'Add Warehouse';
  document.getElementById('wh-edit-idx').value = isEdit ? idx : -1;
  if (isEdit) {
    var w = DATA.warehouses[idx];
    document.getElementById('wh-dc').value = w.dc_code || '';
    document.getElementById('wh-name').value = w.name;
    document.getElementById('wh-address').value = w.address || '';
    document.getElementById('wh-city').value = w.city;
    document.getElementById('wh-state').value = w.state;
    document.getElementById('wh-zip').value = w.zip;
    document.getElementById('wh-group').value = w.group;
    document.getElementById('wh-lat').value = w.lat || '';
    document.getElementById('wh-lon').value = w.lon || '';
  } else {
    ['wh-dc','wh-name','wh-address','wh-city','wh-state','wh-zip','wh-lat','wh-lon'].forEach(function(id) {
      document.getElementById(id).value = '';
    });
    document.getElementById('wh-group').value = 'MS-WH';
  }
  new bootstrap.Modal(document.getElementById('whModal')).show();
}
function editWh(idx) { showWhModal(idx); }

function saveWh() {
  var idx = parseInt(document.getElementById('wh-edit-idx').value);
  var wh = {
    dc_code: document.getElementById('wh-dc').value.trim().toUpperCase(),
    name: document.getElementById('wh-name').value.trim(),
    address: document.getElementById('wh-address').value.trim(),
    city: document.getElementById('wh-city').value.trim(),
    state: document.getElementById('wh-state').value.trim(),
    zip: document.getElementById('wh-zip').value.trim(),
    group: document.getElementById('wh-group').value,
    lat: parseFloat(document.getElementById('wh-lat').value) || 0,
    lon: parseFloat(document.getElementById('wh-lon').value) || 0,
  };
  if (!wh.name) { alert('Name is required'); return; }

  var url = idx >= 0 ? '/api/warehouse/' + idx : '/api/warehouse';
  var method = idx >= 0 ? 'PUT' : 'POST';
  fetch(url, {method: method, headers: {'Content-Type':'application/json'}, body: JSON.stringify(wh)})
    .then(r => r.json()).then(d => {
      if (d.success) {
        bootstrap.Modal.getInstance(document.getElementById('whModal')).hide();
        loadData();
      } else { alert(d.error || 'Failed'); }
    });
}

function deleteWh(idx) {
  var w = DATA.warehouses[idx];
  if (!confirm(w.name + ' 삭제하시겠습니까?')) return;
  fetch('/api/warehouse/' + idx, {method:'DELETE'}).then(r => r.json()).then(d => {
    if (d.success) loadData();
    else alert(d.error || 'Failed');
  });
}

// ---- Schedules ----
function renderSchedules() {
  var tbody = document.getElementById('sched-tbody');
  var scheds = DATA.receiving_schedules || {};
  // Build code->warehouse usage map
  var usage = {};
  for (var i = 0; i < DATA.warehouses.length; i++) {
    var rc = DATA.warehouses[i].receiving_code;
    var dc = DATA.warehouses[i].dc_code || '';
    if (!usage[rc]) usage[rc] = [];
    usage[rc].push((dc ? '[' + dc + '] ' : '') + DATA.warehouses[i].name);
  }
  var html = '';
  var codes = Object.keys(scheds).sort();
  for (var c = 0; c < codes.length; c++) {
    var code = codes[c];
    var wins = scheds[code];
    var desc = wins.map(function(w) { return w[0] + ' ' + w[1] + '-' + w[2]; }).join(' / ');
    var usedBy = (usage[code] || []).join(', ') || '<span class="text-muted">-</span>';
    html += '<tr><td><strong>' + code + '</strong></td><td>' + desc + '</td><td class="small">' + usedBy + '</td>' +
      '<td><button class="btn btn-sm btn-outline-primary py-0 px-1 me-1" onclick="editSched(\'' + code + '\')"><i class="bi bi-pencil"></i></button>' +
      '<button class="btn btn-sm btn-outline-danger py-0 px-1" onclick="deleteSched(\'' + code + '\')"><i class="bi bi-trash"></i></button></td></tr>';
  }
  tbody.innerHTML = html;
}

function addSchedWindow(day, start, end) {
  var div = document.getElementById('sched-windows');
  var row = document.createElement('div');
  row.className = 'input-group input-group-sm mb-1 sched-win';
  row.innerHTML = '<input class="form-control" placeholder="MON-FRI" value="' + (day||'') + '">' +
    '<input class="form-control" placeholder="6AM" value="' + (start||'') + '">' +
    '<input class="form-control" placeholder="1PM" value="' + (end||'') + '">' +
    '<button class="btn btn-outline-danger" onclick="this.parentElement.remove()"><i class="bi bi-x"></i></button>';
  div.appendChild(row);
}

function showSchedModal(code) {
  var isEdit = !!code;
  document.getElementById('schedModalTitle').textContent = isEdit ? 'Edit Schedule: ' + code : 'Add Schedule';
  document.getElementById('sched-edit-code').value = code || '';
  document.getElementById('sched-code').value = code || '';
  document.getElementById('sched-code').disabled = isEdit;
  // Clear old windows
  document.querySelectorAll('.sched-win').forEach(function(el) { el.remove(); });
  if (isEdit && DATA.receiving_schedules[code]) {
    DATA.receiving_schedules[code].forEach(function(w) { addSchedWindow(w[0], w[1], w[2]); });
  } else {
    addSchedWindow();
  }
  new bootstrap.Modal(document.getElementById('schedModal')).show();
}
function editSched(code) { showSchedModal(code); }

function saveSched() {
  var origCode = document.getElementById('sched-edit-code').value;
  var code = document.getElementById('sched-code').value.trim().toUpperCase();
  if (!code) { alert('Code is required'); return; }
  var windows = [];
  document.querySelectorAll('.sched-win').forEach(function(row) {
    var inputs = row.querySelectorAll('input');
    if (inputs[0].value.trim()) {
      windows.push([inputs[0].value.trim(), inputs[1].value.trim(), inputs[2].value.trim()]);
    }
  });
  if (windows.length === 0) { alert('At least one window required'); return; }

  fetch('/api/schedule/' + code, {method:'PUT', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({windows: windows, orig_code: origCode})
  }).then(r => r.json()).then(d => {
    if (d.success) {
      bootstrap.Modal.getInstance(document.getElementById('schedModal')).hide();
      loadData();
    } else { alert(d.error || 'Failed'); }
  });
}

function deleteSched(code) {
  if (!confirm('Schedule ' + code + ' 삭제하시겠습니까?')) return;
  fetch('/api/schedule/' + code, {method:'DELETE'}).then(r => r.json()).then(d => {
    if (d.success) loadData();
    else alert(d.error || 'Failed');
  });
}

// ---- Origins ----
function renderOrigins() {
  var tbody = document.getElementById('origin-tbody');
  var origins = DATA.origins || {};
  var html = '';
  var names = Object.keys(origins).sort();
  for (var i = 0; i < names.length; i++) {
    var n = names[i];
    var o = origins[n];
    html += '<tr><td>' + n + '</td><td>' + o.lat + '</td><td>' + o.lon + '</td>' +
      '<td><button class="btn btn-sm btn-outline-primary py-0 px-1 me-1" onclick="editOrigin(\'' + n + '\')"><i class="bi bi-pencil"></i></button>' +
      '<button class="btn btn-sm btn-outline-danger py-0 px-1" onclick="deleteOrigin(\'' + n + '\')"><i class="bi bi-trash"></i></button></td></tr>';
  }
  tbody.innerHTML = html;
}

function showOriginModal(name) {
  var isEdit = !!name;
  document.getElementById('originModalTitle').textContent = isEdit ? 'Edit Origin' : 'Add Origin';
  document.getElementById('origin-edit-name').value = name || '';
  if (isEdit && DATA.origins[name]) {
    document.getElementById('origin-name').value = name;
    document.getElementById('origin-lat').value = DATA.origins[name].lat;
    document.getElementById('origin-lon').value = DATA.origins[name].lon;
  } else {
    document.getElementById('origin-name').value = '';
    document.getElementById('origin-lat').value = '';
    document.getElementById('origin-lon').value = '';
  }
  new bootstrap.Modal(document.getElementById('originModal')).show();
}
function editOrigin(name) { showOriginModal(name); }

function saveOrigin() {
  var origName = document.getElementById('origin-edit-name').value;
  var name = document.getElementById('origin-name').value.trim();
  var lat = parseFloat(document.getElementById('origin-lat').value);
  var lon = parseFloat(document.getElementById('origin-lon').value);
  if (!name || isNaN(lat) || isNaN(lon)) { alert('All fields required'); return; }

  fetch('/api/origin', {method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({name: name, lat: lat, lon: lon, orig_name: origName})
  }).then(r => r.json()).then(d => {
    if (d.success) {
      bootstrap.Modal.getInstance(document.getElementById('originModal')).hide();
      loadData();
    } else { alert(d.error || 'Failed'); }
  });
}

function deleteOrigin(name) {
  if (!confirm(name + ' 삭제하시겠습니까?')) return;
  fetch('/api/origin/' + encodeURIComponent(name), {method:'DELETE'}).then(r => r.json()).then(d => {
    if (d.success) loadData();
    else alert(d.error || 'Failed');
  });
}

async function uploadRAG(input) {
  if (!input.files || !input.files[0]) return;
  var formData = new FormData();
  formData.append('file', input.files[0]);
  document.getElementById('rag-status').innerHTML = '<div class="alert alert-info"><i class="bi bi-hourglass-split"></i> 파싱 중...</div>';
  try {
    var resp = await fetch('/api/rag-preview', {method: 'POST', body: formData});
    var data = await resp.json();
    if (data.success) {
      window._ragPreview = data.preview;
      var h = '<div class="alert alert-warning"><b>' + data.total_routes + '개 루트, ' + data.total_pos + '개 PO</b> — 확인 후 학습 버튼을 누르세요.</div>';
      h += '<div class="table-responsive"><table class="table table-sm table-bordered small">';
      h += '<thead class="table-dark"><tr><th>Route</th><th>Group</th><th>DC</th><th>PO</th><th>Qty</th><th>Loading</th><th>Pickup</th><th>APPT</th></tr></thead><tbody>';
      data.preview.forEach(function(r) {
        var grpBadge = '<span class="badge '+(r.group==='MS'||r.group==='MS-WH'?'bg-danger':'bg-primary')+'">' + r.group + '</span>';
        var routePickup = r.stops.find(function(s){return s.pickup;});
        var pickupStr = routePickup ? routePickup.pickup : '';
        r.stops.forEach(function(s, si) {
          h += '<tr' + (si===0?' style="border-top:2px solid #000;"':'') + '>';
          h += '<td>' + (si===0?'<b>R'+r.route+'</b>':'') + '</td>';
          h += '<td>' + grpBadge + '</td>';
          h += '<td><b>' + s.dc + '</b></td><td>' + s.po + '</td><td>' + s.qty + '</td>';
          var apptDisplay = '';
          if (s.appt_history && s.appt_history.length > 1) {
            apptDisplay = s.appt_history[0] + ' <span class="badge bg-warning text-dark">' + s.appt_history.length + '회 변경</span><br><small class="text-muted">최종: ' + s.appt_history[s.appt_history.length-1] + '</small>';
          } else {
            apptDisplay = s.appt || '';
          }
          h += '<td>' + (s.loading||'-') + '</td><td>' + (si===0?pickupStr:(s.pickup||'')) + '</td><td>' + apptDisplay + '</td></tr>';
        });
      });
      h += '</tbody></table></div>';
      h += '<button class="btn btn-success" onclick="confirmRAG()"><i class="bi bi-check-lg"></i> 학습 확인</button> ';
      h += '<button class="btn btn-outline-secondary" onclick="cancelRAG()">취소</button>';
      document.getElementById('rag-status').innerHTML = h;
    } else {
      document.getElementById('rag-status').innerHTML = '<div class="alert alert-danger">' + (data.error || 'Error') + '</div>';
    }
  } catch(e) {
    document.getElementById('rag-status').innerHTML = '<div class="alert alert-danger">' + e.message + '</div>';
  }
  input.value = '';
}

async function confirmRAG() {
  if (!window._ragPreview) return;
  document.getElementById('rag-status').innerHTML = '<div class="alert alert-info"><i class="bi bi-hourglass-split"></i> 학습 저장 중...</div>';
  try {
    var resp = await fetch('/api/rag-confirm', {method: 'POST', headers: {'Content-Type':'application/json'}, body: JSON.stringify({routes: window._ragPreview})});
    var data = await resp.json();
    if (data.success) {
      document.getElementById('rag-status').innerHTML = '<div class="alert alert-success"><i class="bi bi-check-circle"></i> ' + data.message + '</div>';
      window._ragPreview = null;
      loadData();
      loadRAGHistory();
    } else {
      document.getElementById('rag-status').innerHTML = '<div class="alert alert-danger">' + (data.error||'Error') + '</div>';
    }
  } catch(e) {
    document.getElementById('rag-status').innerHTML = '<div class="alert alert-danger">' + e.message + '</div>';
  }
}

function cancelRAG() {
  window._ragPreview = null;
  document.getElementById('rag-status').innerHTML = '';
}

async function loadRAGHistory() {
  try {
    var resp = await fetch('/api/rag-history');
    var data = await resp.json();
    var entries = data.entries || [];
    if (entries.length === 0) {
      document.getElementById('rag-history').innerHTML = '<p class="text-muted small">학습된 배차 데이터가 없습니다.</p>';
      return;
    }
    var h = '<table class="table table-sm small mt-3"><thead><tr><th>Date</th><th>Routes</th><th>POs</th></tr></thead><tbody>';
    entries.forEach(function(e) {
      var routes = e.routes.map(function(r) { return r.dcs.join('+') + '(' + r.total_qty + ')'; }).join(' | ');
      h += '<tr><td>' + e.date + '</td><td>' + routes + '</td><td>' + e.total_pos + '</td></tr>';
    });
    h += '</tbody></table>';
    document.getElementById('rag-history').innerHTML = h;
  } catch(e) {}
}

loadData();
loadRAGHistory();
</script>
</body>
</html>
"""


@app.route("/warehouses")
def warehouses_page():
    return render_template_string(WH_MGMT_PAGE)


@app.route("/api/warehouse-data", methods=["GET"])
def api_warehouse_data():
    """Return full warehouse data + DC map for management UI."""
    data = _load_warehouse_data()
    data["dc_map"] = DC_CODE_MAP
    return jsonify(data)


@app.route("/api/warehouse", methods=["POST"])
def api_add_warehouse():
    """Add a new warehouse."""
    wh = request.get_json(force=True)
    data = _load_warehouse_data()
    entry = {
        "dc_code": wh.get("dc_code", ""),
        "name": wh["name"],
        "address": wh.get("address", ""),
        "city": wh.get("city", ""),
        "state": wh.get("state", ""),
        "zip": wh.get("zip", ""),
        "group": wh.get("group", "MS-WH"),
        "receiving_code": wh.get("dc_code", ""),
        "lat": wh.get("lat", 0),
        "lon": wh.get("lon", 0),
    }
    data["warehouses"].append(entry)
    _save_warehouse_data(data)
    _reload_all()
    return jsonify({"success": True})


@app.route("/api/warehouse/<int:idx>", methods=["PUT"])
def api_update_warehouse(idx):
    """Update warehouse at index."""
    wh = request.get_json(force=True)
    data = _load_warehouse_data()
    if idx < 0 or idx >= len(data["warehouses"]):
        return jsonify({"success": False, "error": "Invalid index"}), 400
    data["warehouses"][idx] = {
        "dc_code": wh.get("dc_code", ""),
        "name": wh["name"],
        "address": wh.get("address", ""),
        "city": wh.get("city", ""),
        "state": wh.get("state", ""),
        "zip": wh.get("zip", ""),
        "group": wh.get("group", "MS-WH"),
        "receiving_code": wh.get("dc_code", ""),
        "lat": wh.get("lat", 0),
        "lon": wh.get("lon", 0),
    }
    _save_warehouse_data(data)
    _reload_all()
    return jsonify({"success": True})


@app.route("/api/warehouse/<int:idx>", methods=["DELETE"])
def api_delete_warehouse(idx):
    """Delete warehouse at index."""
    data = _load_warehouse_data()
    if idx < 0 or idx >= len(data["warehouses"]):
        return jsonify({"success": False, "error": "Invalid index"}), 400
    data["warehouses"].pop(idx)
    _save_warehouse_data(data)
    _reload_all()
    return jsonify({"success": True})


@app.route("/api/schedule/<code>", methods=["PUT"])
def api_update_schedule(code):
    """Add or update a receiving schedule."""
    req = request.get_json(force=True)
    data = _load_warehouse_data()
    windows = req.get("windows", [])
    data["receiving_schedules"][code] = windows
    _save_warehouse_data(data)
    _reload_all()
    return jsonify({"success": True})


@app.route("/api/schedule/<code>", methods=["DELETE"])
def api_delete_schedule(code):
    """Delete a receiving schedule."""
    data = _load_warehouse_data()
    if code in data["receiving_schedules"]:
        del data["receiving_schedules"][code]
    _save_warehouse_data(data)
    _reload_all()
    return jsonify({"success": True})


@app.route("/api/origin", methods=["POST"])
def api_save_origin():
    """Add or update an origin."""
    req = request.get_json(force=True)
    data = _load_warehouse_data()
    orig_name = req.get("orig_name", "")
    # Remove old if renaming
    if orig_name and orig_name != req["name"] and orig_name in data["origins"]:
        del data["origins"][orig_name]
    data["origins"][req["name"]] = {"lat": req["lat"], "lon": req["lon"]}
    _save_warehouse_data(data)
    _reload_all()
    return jsonify({"success": True})


@app.route("/api/origin/<path:name>", methods=["DELETE"])
def api_delete_origin(name):
    """Delete an origin."""
    data = _load_warehouse_data()
    if name in data["origins"]:
        del data["origins"][name]
    _save_warehouse_data(data)
    _reload_all()
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# Cache Management Page
# ---------------------------------------------------------------------------
CACHE_PAGE = r"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Distance Cache - Innofoods Route Optimizer</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css" rel="stylesheet">
<style>
  :root { --cj-blue: #003d7a;
  .modal-backdrop { z-index: 10005 !important; }
  .modal { z-index: 10010 !important; } }
  body { background: #f4f6f9; font-size: 0.92rem; }
  .navbar { background: var(--cj-blue) !important; }
  .card { border: none; box-shadow: 0 1px 4px rgba(0,0,0,.08); }
  .btn-cj { background: var(--cj-blue); color: #fff; border: none; }
  .btn-cj:hover { background: #00274d; color: #fff; }
  #log { height: 250px; overflow-y: auto; background: #1e1e1e; color: #0f0; padding: 10px; border-radius: 6px; font-family: monospace; font-size: 0.82rem; }
</style>
</head>
<body>
<nav class="navbar navbar-dark mb-4">
  <div class="container-fluid">
    <span class="navbar-brand mb-0 h1"><i class="bi bi-database"></i> Distance Cache Management</span>
    <a href="/" class="btn btn-outline-light btn-sm"><i class="bi bi-arrow-left"></i> Route Optimizer</a>
  </div>
</nav>

<div class="container" style="max-width:1100px;">

  <!-- Status Card -->
  <div class="card p-3 mb-3">
    <h6><i class="bi bi-info-circle"></i> Cache Status (캐시 현황)</h6>
    <div class="row text-center" id="status-row">
      <div class="col-md-3"><div class="fs-4 fw-bold" id="st-total">-</div><div class="text-muted small">Cached Pairs (저장됨)</div></div>
      <div class="col-md-3"><div class="fs-4 fw-bold" id="st-required">-</div><div class="text-muted small">Required Pairs (필요)</div></div>
      <div class="col-md-3"><div class="fs-4 fw-bold text-danger" id="st-missing">-</div><div class="text-muted small">Missing (미계산)</div></div>
      <div class="col-md-3"><div class="fs-4 fw-bold" id="st-points">-</div><div class="text-muted small">Total Points (지점수)</div></div>
    </div>
  </div>

  <!-- Actions -->
  <div class="card p-3 mb-3">
    <h6><i class="bi bi-gear"></i> Actions</h6>
    <div class="d-flex gap-2 flex-wrap">
      <button class="btn btn-cj" onclick="preCalc(false)"><i class="bi bi-play-fill"></i> Calculate Missing (미계산분 계산)</button>
      <button class="btn btn-warning" onclick="preCalc(true)"><i class="bi bi-arrow-clockwise"></i> Recalculate All (전체 재계산)</button>
      <button class="btn btn-outline-secondary" onclick="loadStatus()"><i class="bi bi-arrow-repeat"></i> Refresh</button>
    </div>
    <div id="log" class="mt-3" style="display:none;"></div>
  </div>

  <!-- Distance Table -->
  <div class="card p-3 mb-3">
    <div class="d-flex justify-content-between align-items-center mb-2">
      <h6 class="mb-0"><i class="bi bi-table"></i> Distance Lookup (거리 조회)</h6>
      <div>
        <select id="filter-from" class="form-select form-select-sm d-inline-block" style="width:250px;" onchange="filterTable()">
          <option value="">All Origins (전체 출발지)</option>
        </select>
      </div>
    </div>
    <div class="table-responsive" style="max-height:500px; overflow-y:auto;">
      <table class="table table-sm table-bordered table-striped small" id="dist-table">
        <thead class="table-dark sticky-top">
          <tr><th>From</th><th>To</th><th>Distance (km)</th><th>Status</th></tr>
        </thead>
        <tbody id="dist-tbody"></tbody>
      </table>
    </div>
  </div>

</div>

<script>
var allPairs = [];

function loadStatus() {
  fetch('/api/cache/status').then(r => r.json()).then(data => {
    document.getElementById('st-total').textContent = data.cached;
    document.getElementById('st-required').textContent = data.required;
    document.getElementById('st-missing').textContent = data.missing;
    document.getElementById('st-points').textContent = data.points;

    allPairs = data.pairs || [];
    populateFilter(data.point_names || []);
    renderTable(allPairs);
  });
}

function populateFilter(names) {
  var sel = document.getElementById('filter-from');
  var cur = sel.value;
  sel.innerHTML = '<option value="">All (전체)</option>';
  for (var i = 0; i < names.length; i++) {
    sel.innerHTML += '<option value="' + names[i] + '">' + names[i] + '</option>';
  }
  sel.value = cur;
}

function filterTable() {
  var f = document.getElementById('filter-from').value;
  if (!f) { renderTable(allPairs); return; }
  var filtered = allPairs.filter(function(p) { return p.from === f || p.to === f; });
  renderTable(filtered);
}

function renderTable(pairs) {
  var tbody = document.getElementById('dist-tbody');
  var html = '';
  for (var i = 0; i < pairs.length; i++) {
    var p = pairs[i];
    var cls = p.cached ? '' : 'table-warning';
    var dist = p.cached ? p.distance.toLocaleString(undefined, {maximumFractionDigits:1}) : '-';
    var st = p.cached ? '<span class="text-success">Cached</span>' : '<span class="text-danger">Missing</span>';
    html += '<tr class="' + cls + '"><td>' + p.from + '</td><td>' + p.to + '</td><td>' + dist + '</td><td>' + st + '</td></tr>';
  }
  tbody.innerHTML = html;
}

function preCalc(recalcAll) {
  var logEl = document.getElementById('log');
  logEl.style.display = 'block';
  logEl.innerHTML = '> Starting ' + (recalcAll ? 'full recalculation' : 'missing calculation') + '...\n';

  fetch('/api/cache/precalc', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({recalc_all: recalcAll})
  }).then(r => r.json()).then(data => {
    logEl.innerHTML += '> Computed: ' + data.computed + ' pairs\n';
    logEl.innerHTML += '> Failed: ' + data.failed + ' pairs\n';
    logEl.innerHTML += '> Total cache: ' + data.total_cache + '\n';
    logEl.innerHTML += '> Time: ' + data.elapsed + ' seconds\n';
    if (data.errors && data.errors.length > 0) {
      for (var i = 0; i < data.errors.length; i++) {
        logEl.innerHTML += '> ERROR: ' + data.errors[i] + '\n';
      }
    }
    logEl.innerHTML += '> Done!\n';
    logEl.scrollTop = logEl.scrollHeight;
    loadStatus();
  }).catch(function(e) {
    logEl.innerHTML += '> FAILED: ' + e.message + '\n';
  });
}

loadStatus();
</script>
</body>
</html>
"""


@app.route("/cache")
def cache_page():
    return render_template_string(CACHE_PAGE)


@app.route("/api/cache/status", methods=["GET"])
def api_cache_status():
    """Return cache status: how many pairs cached, missing, etc."""
    cache = load_cache()

    # All points
    points = {}
    for name, coord in ORIGIN_COORDS.items():
        points[name] = coord
    for wh in WAREHOUSES:
        if wh["name"] in COORDS:
            points[wh["name"]] = COORDS[wh["name"]]

    names = sorted(points.keys())
    pairs = []
    cached_count = 0
    missing_count = 0

    for i in range(len(names)):
        for j in range(i + 1, len(names)):
            c1 = points[names[i]]
            c2 = points[names[j]]
            key = f"{c1[0]:.4f},{c1[1]:.4f}|{c2[0]:.4f},{c2[1]:.4f}"
            key_r = f"{c2[0]:.4f},{c2[1]:.4f}|{c1[0]:.4f},{c1[1]:.4f}"
            val = cache.get(key) or cache.get(key_r)
            is_cached = val is not None
            if is_cached:
                cached_count += 1
                if isinstance(val, list):
                    dist_val = val[0]
                    dur_val = val[1]
                else:
                    dist_val = val
                    dur_val = round(val / 80, 1)
            else:
                missing_count += 1
                dist_val = 0
                dur_val = 0
            pairs.append({
                "from": names[i],
                "to": names[j],
                "cached": is_cached,
                "distance": round(dist_val, 1),
                "hours": round(dur_val, 1),
            })

    # Sort: missing first, then by from/to
    pairs.sort(key=lambda p: (p["cached"], p["from"], p["to"]))

    return jsonify({
        "points": len(names),
        "point_names": names,
        "required": len(names) * (len(names) - 1) // 2,
        "cached": cached_count,
        "missing": missing_count,
        "pairs": pairs,
    })


@app.route("/api/cache/precalc", methods=["POST"])
def api_cache_precalc():
    """Pre-calculate missing (or all) pairwise distances."""
    import time as _time
    import requests as _requests

    data = request.get_json(force=True) or {}
    recalc_all = data.get("recalc_all", False)

    cache = load_cache()
    start = _time.time()

    points = {}
    for name, coord in ORIGIN_COORDS.items():
        points[name] = coord
    for wh in WAREHOUSES:
        if wh["name"] in COORDS:
            points[wh["name"]] = COORDS[wh["name"]]

    names = sorted(points.keys())
    computed = 0
    failed = 0
    errors = []

    for i in range(len(names)):
        for j in range(i + 1, len(names)):
            c1 = points[names[i]]
            c2 = points[names[j]]
            key = f"{c1[0]:.4f},{c1[1]:.4f}|{c2[0]:.4f},{c2[1]:.4f}"
            key_r = f"{c2[0]:.4f},{c2[1]:.4f}|{c1[0]:.4f},{c1[1]:.4f}"

            if not recalc_all and (key in cache or key_r in cache):
                continue

            # Remove old keys if recalculating
            if recalc_all:
                cache.pop(key, None)
                cache.pop(key_r, None)

            coord_str = f"{c1[1]},{c1[0]};{c2[1]},{c2[0]}"
            url = f"https://router.project-osrm.org/route/v1/driving/{coord_str}?overview=false"

            ok = False
            for attempt in range(3):
                try:
                    _time.sleep(0.3)
                    resp = _requests.get(url, timeout=15)
                    if resp.status_code == 429:
                        _time.sleep(2 * (attempt + 1))
                        continue
                    rdata = resp.json()
                    if rdata.get("code") == "Ok" and rdata.get("routes"):
                        dist_km = rdata["routes"][0]["distance"] / 1000.0
                        cache[key] = dist_km
                        computed += 1
                        ok = True
                        break
                except Exception as e:
                    _time.sleep(1)

            if not ok:
                failed += 1
                errors.append(f"{names[i]} -> {names[j]}")

            # Save periodically
            if computed % 50 == 0 and computed > 0:
                save_cache(cache)

    save_cache(cache)
    elapsed = round(_time.time() - start, 1)

    return jsonify({
        "computed": computed,
        "failed": failed,
        "total_cache": len(cache),
        "elapsed": elapsed,
        "errors": errors,
    })


# ---------------------------------------------------------------------------
# Learned Preferences Page
# ---------------------------------------------------------------------------

LEARNED_PAGE = r"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>학습 이력 - Route Optimizer</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css" rel="stylesheet">
<style>
  :root { --cj-blue: #003d7a;
  .modal-backdrop { z-index: 10005 !important; }
  .modal { z-index: 10010 !important; } }
  body { background: linear-gradient(135deg, #f0f2f5 0%, #e2e8f0 100%); font-size: 0.92rem; }
  .navbar { background: linear-gradient(90deg, #001f3f 0%, var(--cj-blue) 50%, #004e9a 100%) !important; }
  .card { border: none; box-shadow: 0 2px 12px rgba(0,0,0,.06); border-radius: 12px; }
  .pair-badge { font-size: 1rem; padding: 6px 14px; margin: 3px; display: inline-block; }
  .score-bar { height: 8px; border-radius: 4px; background: #e9ecef; overflow: hidden; }
  .score-fill { height: 100%; border-radius: 4px; background: linear-gradient(90deg, #2ecc71, #27ae60); }
</style>
</head>
<body>
<nav class="navbar navbar-dark mb-4">
  <div class="container-fluid">
    <a class="navbar-brand" href="/"><i class="bi bi-truck"></i> Innofoods Route Optimizer</a>
    <div class="d-flex gap-2">
      <a href="/" class="btn btn-outline-light btn-sm"><i class="bi bi-arrow-left"></i> Back</a>
      <a href="/warehouses" class="btn btn-outline-light btn-sm"><i class="bi bi-building"></i> Warehouse</a>
    </div>
  </div>
</nav>

<div class="container" style="max-width:1000px;">
  <h4 class="mb-4"><i class="bi bi-mortarboard"></i> 학습된 DC 페어링 & 이력</h4>

  <!-- Pair Scores -->
  <div class="card p-4 mb-4">
    <h5><i class="bi bi-link-45deg"></i> DC 페어링 점수</h5>
    <p class="text-muted small">자주 같이 묶인 DC 조합일수록 점수가 높습니다. 점수가 높을수록 다음 최적화에서 우선 선택됩니다.</p>
    <div id="pair-scores"></div>
    <div class="mt-3">
      <button class="btn btn-sm btn-outline-danger" onclick="resetPrefs()"><i class="bi bi-trash"></i> 학습 초기화</button>
    </div>
  </div>

  <!-- Saved Results -->
  <div class="card p-4 mb-4">
    <h5><i class="bi bi-clock-history"></i> 학습 이력 (최근 50건)</h5>
    <div id="saved-results"></div>
  </div>

  <!-- RAG Training Data -->
  <div class="card p-4 mb-4">
    <h5><i class="bi bi-database-add"></i> 과거 배차 데이터 학습 (RAG)</h5>
    <p class="text-muted small">과거 실제 배차 엑셀을 업로드하면 AI가 패턴을 학습합니다. 같은 ROUTE 번호 = 같은 트럭.</p>
    <div class="d-flex gap-2 mb-3">
      <a href="/api/rag-template.xlsx" class="btn btn-sm btn-outline-success"><i class="bi bi-download"></i> 템플릿 다운로드</a>
      <button class="btn btn-sm btn-outline-warning" onclick="document.getElementById('rag-upload').click()"><i class="bi bi-upload"></i> 배차 데이터 업로드</button>
      <input type="file" id="rag-upload" accept=".xlsx,.xls" style="display:none" onchange="uploadRAG(this)">
    </div>
    <div id="rag-status"></div>
    <div id="rag-history"></div>
  </div>
</div>

<script>
async function loadData() {
  var resp = await fetch('/api/preferences');
  var data = await resp.json();

  // Pair scores
  var pairs = data.pair_scores || {};
  var keys = Object.keys(pairs).sort((a,b) => pairs[b] - pairs[a]);
  var maxScore = keys.length > 0 ? Math.max(...Object.values(pairs)) : 1;

  var html = '';
  if (keys.length === 0) {
    html = '<p class="text-muted">아직 학습된 패턴이 없습니다. 결과 페이지에서 "이 결과 학습" 버튼을 눌러주세요.</p>';
  } else {
    html = '<table class="table table-sm"><thead><tr><th>DC Pair</th><th>Score</th><th style="width:200px;">Strength</th><th></th></tr></thead><tbody>';
    keys.forEach(function(k) {
      var score = pairs[k];
      var pct = Math.round((score / Math.max(maxScore, 1)) * 100);
      var dcs = k.split('|');
      html += '<tr>';
      html += '<td><span class="badge bg-primary pair-badge">' + dcs[0] + '</span> + <span class="badge bg-primary pair-badge">' + dcs[1] + '</span></td>';
      html += '<td class="fw-bold">' + score + '</td>';
      html += '<td><div class="score-bar"><div class="score-fill" style="width:' + pct + '%;"></div></div></td>';
      html += '<td><button class="btn btn-sm btn-outline-danger py-0 px-1" onclick="deletePair(\'' + k + '\')"><i class="bi bi-x"></i></button></td>';
      html += '</tr>';
    });
    html += '</tbody></table>';
    html += '<p class="small text-muted">Bonus: 점수 x $100 (최대 $1,000/pair)</p>';
  }
  document.getElementById('pair-scores').innerHTML = html;

  // Saved results
  var saved = (data.saved_results || []).slice().reverse();
  var rhtml = '';
  if (saved.length === 0) {
    rhtml = '<p class="text-muted">이력이 없습니다.</p>';
  } else {
    rhtml = '<div class="table-responsive"><table class="table table-sm table-striped small">';
    rhtml += '<thead><tr><th>#</th><th>Date</th><th>Group</th><th>Routes</th><th>Total Cost</th></tr></thead><tbody>';
    saved.forEach(function(s, idx) {
      var routeStr = s.routes.map(function(r) {
        return r.dcs.join('+');
      }).join(' | ');
      rhtml += '<tr>';
      rhtml += '<td>' + (saved.length - idx) + '</td>';
      rhtml += '<td>' + s.timestamp + '</td>';
      rhtml += '<td><span class="badge ' + (s.group === 'MS-WH' ? 'bg-danger' : 'bg-primary') + '">' + s.group + '</span></td>';
      rhtml += '<td style="font-size:0.8rem;">' + routeStr + '</td>';
      rhtml += '<td>$' + (s.total_cost || 0).toLocaleString() + '</td>';
      rhtml += '</tr>';
    });
    rhtml += '</tbody></table></div>';
  }
  document.getElementById('saved-results').innerHTML = rhtml;
}

async function deletePair(key) {
  if (!confirm(key.replace('|', ' + ') + ' 페어링을 삭제하시겠습니까?')) return;
  await fetch('/api/preferences/delete-pair', {method: 'POST', headers: {'Content-Type':'application/json'}, body: JSON.stringify({pair: key})});
  loadData();
}

async function resetPrefs() {
  if (!confirm('학습 데이터를 모두 초기화하시겠습니까?')) return;
  await fetch('/api/preferences/reset', {method: 'POST'});
  loadData();
}

async function uploadRAG(input) {
  if (!input.files || !input.files[0]) return;
  var formData = new FormData();
  formData.append('file', input.files[0]);
  document.getElementById('rag-status').innerHTML = '<div class="alert alert-info"><i class="bi bi-hourglass-split"></i> 파싱 중...</div>';
  try {
    var resp = await fetch('/api/rag-preview', {method: 'POST', body: formData});
    var data = await resp.json();
    if (data.success) {
      window._ragPreview = data.preview;
      var h = '<div class="alert alert-warning"><b>' + data.total_routes + '개 루트, ' + data.total_pos + '개 PO</b> — 확인 후 학습 버튼을 누르세요.</div>';
      h += '<div class="table-responsive"><table class="table table-sm table-bordered small">';
      h += '<thead class="table-dark"><tr><th>Route</th><th>Group</th><th>DC</th><th>PO</th><th>Qty</th><th>Loading</th><th>Pickup</th><th>APPT</th></tr></thead><tbody>';
      data.preview.forEach(function(r) {
        var grpBadge = '<span class="badge '+(r.group==='MS'||r.group==='MS-WH'?'bg-danger':'bg-primary')+'">' + r.group + '</span>';
        var routePickup = r.stops.find(function(s){return s.pickup;});
        var pickupStr = routePickup ? routePickup.pickup : '';
        r.stops.forEach(function(s, si) {
          h += '<tr' + (si===0?' style="border-top:2px solid #000;"':'') + '>';
          h += '<td>' + (si===0?'<b>R'+r.route+'</b>':'') + '</td>';
          h += '<td>' + grpBadge + '</td>';
          h += '<td><b>' + s.dc + '</b></td><td>' + s.po + '</td><td>' + s.qty + '</td>';
          var apptDisplay = '';
          if (s.appt_history && s.appt_history.length > 1) {
            apptDisplay = s.appt_history[0] + ' <span class="badge bg-warning text-dark">' + s.appt_history.length + '회 변경</span><br><small class="text-muted">최종: ' + s.appt_history[s.appt_history.length-1] + '</small>';
          } else {
            apptDisplay = s.appt || '';
          }
          h += '<td>' + (s.loading||'-') + '</td><td>' + (si===0?pickupStr:(s.pickup||'')) + '</td><td>' + apptDisplay + '</td></tr>';
        });
      });
      h += '</tbody></table></div>';
      h += '<button class="btn btn-success" onclick="confirmRAG()"><i class="bi bi-check-lg"></i> 학습 확인</button> ';
      h += '<button class="btn btn-outline-secondary" onclick="cancelRAG()">취소</button>';
      document.getElementById('rag-status').innerHTML = h;
    } else {
      document.getElementById('rag-status').innerHTML = '<div class="alert alert-danger">' + (data.error || 'Error') + '</div>';
    }
  } catch(e) {
    document.getElementById('rag-status').innerHTML = '<div class="alert alert-danger">' + e.message + '</div>';
  }
  input.value = '';
}

async function confirmRAG() {
  if (!window._ragPreview) return;
  document.getElementById('rag-status').innerHTML = '<div class="alert alert-info"><i class="bi bi-hourglass-split"></i> 학습 저장 중...</div>';
  try {
    var resp = await fetch('/api/rag-confirm', {method: 'POST', headers: {'Content-Type':'application/json'}, body: JSON.stringify({routes: window._ragPreview})});
    var data = await resp.json();
    if (data.success) {
      document.getElementById('rag-status').innerHTML = '<div class="alert alert-success"><i class="bi bi-check-circle"></i> ' + data.message + '</div>';
      window._ragPreview = null;
      loadData();
      loadRAGHistory();
    } else {
      document.getElementById('rag-status').innerHTML = '<div class="alert alert-danger">' + (data.error||'Error') + '</div>';
    }
  } catch(e) {
    document.getElementById('rag-status').innerHTML = '<div class="alert alert-danger">' + e.message + '</div>';
  }
}

function cancelRAG() {
  window._ragPreview = null;
  document.getElementById('rag-status').innerHTML = '';
}

async function loadRAGHistory() {
  try {
    var resp = await fetch('/api/rag-history');
    var data = await resp.json();
    var entries = data.entries || [];
    if (entries.length === 0) {
      document.getElementById('rag-history').innerHTML = '<p class="text-muted small">학습된 배차 데이터가 없습니다.</p>';
      return;
    }
    var h = '<table class="table table-sm small mt-3"><thead><tr><th>Date</th><th>Routes</th><th>POs</th></tr></thead><tbody>';
    entries.forEach(function(e) {
      var routes = e.routes.map(function(r) { return r.dcs.join('+') + '(' + r.total_qty + ')'; }).join(' | ');
      h += '<tr><td>' + e.date + '</td><td>' + routes + '</td><td>' + e.total_pos + '</td></tr>';
    });
    h += '</tbody></table>';
    document.getElementById('rag-history').innerHTML = h;
  } catch(e) {}
}

loadData();
loadRAGHistory();
</script>
</body>
</html>
"""


@app.route("/api/rag-template.xlsx", methods=["GET"])
def rag_template():
    """Download RAG training data template."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "배차 데이터"

    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="003D7A", end_color="003D7A", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    headers = [("PU from", 10), ("SHIP TO", 45), ("DC", 8), ("PO #", 22),
               ("QTY\n(PALLETS)", 12), ("Loading", 10), ("PICK UP", 18), ("APPT  (Delivery)", 28), ("Hours", 8), ("Due date", 12), ("WAVE", 8)]
    col_letters = 'ABCDEFGHIJK'
    for col, (name, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = hfont; cell.fill = hfill; cell.alignment = halign; cell.border = border
        ws.column_dimensions[col_letters[col-1]].width = width

    samples = [
        ("MS", "MCLANE OHIO\n3200 MCLANE DRIVE\nFINDLAY, OH 45840", "MG", "MG10186484-01", 48, "-", "1/12 @ 2pm", "1/13 @ 8:30PM EST", 6, "1/13", 46),
        ("", "MCLANE CUMBERLAND\n104 MCLANE BLVD\nNICHOLASVILLE, KY 40356", "MK", "MK10087938-01", 20, "-", "1/7 @ 11am", "1/8 @8:00PM EST", 10, "1/13", 46),
        ("", "MCLANE PA\n43 VALLY VIEW BUSINESS PA\nJESSUP, PA 18434", "PA", "PA10083786-01", 20, "Tail", "1/5 @ 3pm", "1/6 @ 7:00AM EST", 7, "1/13", 46),
        ("", "MCLANE MID-ATLANTIC\n56 MCLANE DRIVE\nFredricksburg, VA 22406", "MZ", "MZ10081807-01", 10, "Middle", "", "1/6 @4:00pm EST", 5, "1/13", 46),
        ("", "MCLANE CAROLINA\n7253 NC-4\nBATTLEBORO, NC 27809", "NC", "NC10094703-01", 20, "Nose", "", "1/7 @ 3:30PM EST", 3, "1/13", 46),
        ("", "MCLANE SOUTHEAST\n300 NORTH HWY 29\nATHENS, GA 30601", "SE", "SE10193724-01", 20, "Tail", "1/5 @ 1pm", "1/7 @ 11:00PM EST", 16, "1/13", 46),
        ("", "MCLANE OCALA\n910 NW 50TH AVE\nOCALA, FL 34482", "FE", "FE10187784-01", 10, "Middle", "", "1/8 @6:00PM EST", 6, "1/13", 46),
        ("", "MCLANE SUNEAST\n1818 POINCIANA BLVD\nKISSIMMEE, FL 34758", "ME", "ME10127808-01", 20, "Nose", "", "1/8 @11:00PM EST", 2, "1/13", 46),
        ("L1", "MCLANE NORTHWEST\n9611 45TH AVE SW BLDG 4\nLAKEWOOD, WA 98499", "NW", "NW10086469-01", 4, "Tail", "1/3 @ 12pm", "1/4 @8:00PM PST", 4, "1/13", 46),
        ("", "MCLANE WESTERN\n2100 E KEN PRATT BLVD\nLONGMONT, CO 80504", "MW", "MW10093811-01", 44, "Nose", "", "1/6 @8:00PM MST", 20, "1/13", 46),
    ]
    wrap_align = Alignment(vertical="top", wrap_text=True)
    for r, row_data in enumerate(samples, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            if c == 2:  # SHIP TO - wrap text
                cell.alignment = wrap_align
        ws.row_dimensions[r].height = 45

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=rag_training_template.xlsx"})


@app.route("/api/rag-preview", methods=["POST"])
def rag_preview():
    """Upload and preview parsed dispatch data before learning."""
    from openpyxl import load_workbook
    from io import BytesIO

    if "file" not in request.files:
        return jsonify({"success": False, "error": "No file"})

    f = request.files["file"]
    wb = load_workbook(BytesIO(f.read()), data_only=True)
    ws = wb.active

    headers = [str(cell.value or "").strip().upper() for cell in ws[1]]
    col = {}
    for i, h in enumerate(headers):
        if h == "ROUTE": col["route"] = i
        elif "PU" in h and "FROM" in h: col["pu"] = i
        elif "SHIP" in h and "TO" in h: col["shipto"] = i
        elif h == "DC": col["dc"] = i
        elif "PO" in h and "#" in h: col["po"] = i
        elif "QTY" in h or "PALLET" in h: col["qty"] = i
        elif "PICK" in h: col["pickup"] = i
        elif "APPT" in h or ("DELIVERY" in h and "DUE" not in h): col["appt"] = i
        elif "HOUR" in h or h == "HRS": col["hours"] = i
        elif "DUE" in h: col["due"] = i
        elif "WAVE" in h: col["wave"] = i

    if "dc" not in col:
        return jsonify({"success": False, "error": "DC 컬럼이 필요합니다."})

    # Also detect Loading and Hours columns
    for i, h in enumerate(headers):
        if "LOADING" in h: col["loading"] = i
        elif h == "HOURS" or h == "HRS": col["hours"] = i
        elif "WAVE" in h: col["wave"] = i

    # Parse rows into routes
    # Two modes: 1) ROUTE column exists, or 2) group by PICK UP (new pickup = new truck)
    routes_map = {}
    current_route = 0
    has_route_col = "route" in col

    import sys
    print(f"RAG headers: {headers}", file=sys.stderr)
    print(f"RAG col map: {col}", file=sys.stderr)
    row_count = 0
    skip_count = 0
    last_pu = "MS"  # carry-forward for merged PU FROM cells
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        row_count += 1
        
        dc = str(vals[col["dc"]] or "").strip().upper() if col.get("dc") is not None and col["dc"] < len(vals) else ""
        if not dc:
            skip_count += 1
            if row_count <= 5 or row_count % 10 == 0:
                print(f"RAG row {row_count} SKIP: dc empty, vals={[str(v)[:20] if v else '' for v in vals[:6]]}", file=sys.stderr)
            continue
        
        po = str(vals[col.get("po", 0)] or "").strip() if col.get("po") is not None else ""
        qty = 0
        if col.get("qty") is not None:
            try: qty = int(float(str(vals[col["qty"]] or 0).replace(",","")))
            except: pass
        pu = str(vals[col.get("pu", 0)] or "").strip().upper() if col.get("pu") is not None else ""
        if pu:
            last_pu = pu  # update carry-forward
        else:
            pu = last_pu  # inherit from previous row (merged cell)
        due = str(vals[col.get("due", 0)] or "").strip() if col.get("due") is not None else ""
        appt_raw = str(vals[col.get("appt", 0)] or "").strip() if col.get("appt") is not None else ""
        appt_history = []
        if '->' in appt_raw:
            appt_history = [p.strip() for p in appt_raw.split('->')]
            appt = appt_history[0]  # first value (original plan) for learning
        else:
            appt = appt_raw
        pickup = str(vals[col.get("pickup", 0)] or "").strip() if col.get("pickup") is not None else ""
        loading = str(vals[col.get("loading", 0)] or "").strip() if col.get("loading") is not None else ""

        if has_route_col:
            route_num = str(vals[col["route"]] or "").strip()
            if not route_num: route_num = str(current_route)
        else:
            # No ROUTE column: new PICK UP time = new truck
            if pickup:
                current_route += 1
            route_num = str(current_route)

        if route_num not in routes_map:
            routes_map[route_num] = {"group": pu or "MS", "stops": []}
        routes_map[route_num]["stops"].append({
            "dc": dc, "po": po, "qty": qty, "due": due, 
            "appt": appt, "appt_history": appt_history, "pickup": pickup, "loading": loading
        })
        if pu: routes_map[route_num]["group"] = pu

    # Return preview data, sort numerically
    preview = []
    def route_sort_key(item):
        try: return int(item[0])
        except: return 999
    for rnum, rdata in sorted(routes_map.items(), key=route_sort_key):
        stops = []
        for s in rdata["stops"]:
            stops.append({"dc": s["dc"], "po": s["po"], "qty": s["qty"], "appt": s.get("appt",""), "appt_history": s.get("appt_history",[]), "loading": s.get("loading",""), "pickup": s.get("pickup","")})
        preview.append({"route": rnum, "group": rdata["group"], "stops": stops, "total_qty": sum(s["qty"] for s in rdata["stops"])})

    print(f"RAG parse: {row_count} rows, {skip_count} skipped, {len(routes_map)} routes", file=sys.stderr)
    return jsonify({"success": True, "preview": preview, "total_routes": len(preview), "total_pos": sum(len(r["stops"]) for r in preview)})


@app.route("/api/rag-confirm", methods=["POST"])
def rag_confirm():
    """Confirm and save RAG learning data."""
    from route_optimizer import load_preferences, save_preferences

    data = request.get_json(force=True)
    routes = data.get("routes", [])
    if not routes:
        return jsonify({"success": False, "error": "No routes"})

    prefs = load_preferences()
    if "rag_data" not in prefs: prefs["rag_data"] = []
    pair_scores = prefs.get("pair_scores", {})

    rag_entry = {
        "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "total_pos": sum(len(r["stops"]) for r in routes),
        "routes": []
    }

    for r in routes:
        dcs = [s["dc"] for s in r["stops"]]
        appts = [s.get("appt","") for s in r["stops"] if s.get("appt")]
        appt_changes = [s.get("appt_history",[]) for s in r["stops"] if s.get("appt_history")]
        rag_entry["routes"].append({"route": r["route"], "group": r["group"], "dcs": dcs, "total_qty": r["total_qty"], "appts": appts, "appt_changes": appt_changes})

        for i in range(len(dcs)):
            for j in range(i+1, len(dcs)):
                pair_key = "|".join(sorted([dcs[i], dcs[j]]))
                pair_scores[pair_key] = pair_scores.get(pair_key, 0) + 2

    prefs["pair_scores"] = pair_scores
    prefs["rag_data"].append(rag_entry)
    prefs["rag_data"] = prefs["rag_data"][-100:]
    save_preferences(prefs)

    return jsonify({"success": True, "message": f"{len(routes)}개 루트, {rag_entry['total_pos']}개 PO 학습 완료."})


@app.route("/api/rag-history", methods=["GET"])
def rag_history():
    from route_optimizer import load_preferences
    prefs = load_preferences()
    entries = prefs.get("rag_data", [])
    return jsonify({"entries": entries[-20:]})


@app.route("/learned")
def learned_page():
    return render_template_string(LEARNED_PAGE)


@app.route("/api/preferences/delete-pair", methods=["POST"])
def api_delete_pair():
    from route_optimizer import load_preferences, save_preferences
    data = request.get_json(force=True)
    pair = data.get("pair", "")
    prefs = load_preferences()
    prefs.get("pair_scores", {}).pop(pair, None)
    save_preferences(prefs)
    return jsonify({"success": True})


@app.route("/api/preferences/reset", methods=["POST"])
def api_reset_preferences():
    from route_optimizer import save_preferences
    save_preferences({"pair_scores": {}, "saved_results": []})
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import socket

    port = 8105
    # Check if port is available
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    result = sock.connect_ex(("127.0.0.1", port))
    sock.close()
    if result == 0:
        print(f"WARNING: Port {port} appears to be in use.")
        print(f"  Stop the existing process first, or this app may fail to bind.")
        print(f"  Try: fuser -k {port}/tcp")
        print()

    print(f"Starting Route Optimizer Web on http://0.0.0.0:{port}")
    app.run(host="0.0.0.0", port=port, debug=False)
